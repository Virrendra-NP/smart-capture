import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import io
from datetime import datetime
from PIL import Image, ExifTags
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

def parse_date_from_filename(filename):
    """WhatsApp stores date in filename as IMG-YYYYMMDD-WAxxxx"""
    match = re.search(r'(\d{8})', filename)
    if match:
        d_str = match.group(1)
        try:
            return datetime.strptime(d_str, '%Y%m%d').strftime('%Y-%m-%d (From Filename)')
        except: pass
    return None

# Optional AI Libraries
AI_AVAILABLE = False
try:
    import torch
    import torchvision.transforms as T
    from torchvision.models import mobilenet_v2, MobileNet_V2_Weights
    AI_AVAILABLE = True
except ImportError:
    pass

class ConstructionAI:
    def __init__(self):
        if not AI_AVAILABLE: return
        try:
            weights = MobileNet_V2_Weights.DEFAULT
            self.model = mobilenet_v2(weights=weights)
            self.model.eval()
            self.preprocess = weights.transforms()
            self.categories = weights.meta["categories"]
        except:
            self.model = None

    def recognize(self, img_path):
        if not AI_AVAILABLE or not hasattr(self, 'model') or not self.model:
            return "Site Photo"
        
        try:
            img = Image.open(img_path).convert("RGB")
            batch = self.preprocess(img).unsqueeze(0)
            with torch.no_grad():
                prediction = self.model(batch).squeeze(0).softmax(0)
                class_id = prediction.argmax().item()
                label = self.categories[class_id].lower()
            
            # Mapping logic
            if any(k in label for k in ['tractor', 'rig', 'shovel', 'dirt', 'trench', 'backhoe']):
                return "EARTH WORK"
            if any(k in label for k in ['scaffolding', 'lumber', 'beam', 'pole', 'formwork']):
                return "SHUTTERING WORK"
            if any(k in label for k in ['building', 'house', 'structure', 'wall', 'brick', 'concrete']):
                return "BUILDING WORK"
            if any(k in label for k in ['crane', 'industrial']):
                return "CRANE / HEAVY EQUIP"
            
            return f"Site Activity ({label.title()})"
        except:
            return "General Site Work"

ai_engine = ConstructionAI()

def get_exif_data(image_path):
    try:
        img = Image.open(image_path)
        exif_data = img._getexif()
        if not exif_data: return None
        decoded = {}
        for tag, value in exif_data.items():
            decoded[ExifTags.TAGS.get(tag, tag)] = value
        return decoded
    except: return None

def get_lat_lon(exif_data):
    if not exif_data or 'GPSInfo' not in exif_data: return None, None
    def to_deg(v): return float(v[0]) + (float(v[1]) / 60.0) + (float(v[2]) / 3600.0)
    try:
        gps = exif_data['GPSInfo']
        lat = to_deg(gps[2]); lon = to_deg(gps[4])
        if gps[1] == 'S': lat = -lat
        if gps[3] == 'W': lon = -lon
        return lat, lon
    except: return None, None

def process_photos():
    folder_path = folder_var.get()
    excel_path = excel_var.get()
    
    if not folder_path or not os.path.exists(folder_path):
        messagebox.showerror("Error", "Select photo folder")
        return
        
    transfer_btn.config(state="disabled", text="AI THINKING...")
    root.update()

    try:
        photos = [f for f in os.listdir(folder_path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        if not photos:
            messagebox.showinfo("Info", "No photos found"); transfer_btn.config(state="normal", text="Transfer Photos"); return

        if os.path.exists(excel_path):
            try: wb = load_workbook(excel_path)
            except: 
                messagebox.showerror("Error", "Close Excel before updating"); transfer_btn.config(state="normal", text="Transfer Photos"); return
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames: del wb["Sheet"]

        # UNIQUE SHEET NAME
        base_name = f"Group_{datetime.now().strftime('%d%b_%H%M')}"
        ws = wb.create_sheet(base_name)
        ws.sheet_view.showGridLines = False

        # STYLING
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        label_font = Font(bold=True, size=9)
        border = Border(bottom=Side(style='thin', color="cbd5e1"))

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50

        row_idx = 1
        for i in range(0, len(photos), 2):
            # Row settings
            ws.row_dimensions[row_idx].height = 18   # AI Title
            ws.row_dimensions[row_idx+1].height = 18 # Meta 1
            ws.row_dimensions[row_idx+2].height = 18 # Meta 2
            ws.row_dimensions[row_idx+3].height = 250 # Image

            for j in range(2):
                if i + j < len(photos):
                    p_name = photos[i+j]; p_path = os.path.join(folder_path, p_name); col = j + 1
                    
                    # AI RECOGNITION
                    tag = ai_engine.recognize(p_path)
                    
                    # METADATA RECOVERY
                    exif = get_exif_data(p_path)
                    date_val = None
                    if exif: date_val = exif.get('DateTimeOriginal', exif.get('DateTime'))
                    
                    if not date_val: # Try filename (WhatsApp pattern)
                        date_val = parse_date_from_filename(p_name)
                    
                    if not date_val: # Try file creation date
                        ts = os.path.getmtime(p_path)
                        date_val = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M (Approx)')
                        
                    lat, lon = get_lat_lon(exif)
                    loc_str = f"{round(lat,4)},{round(lon,4)}" if lat else "No GPS (WhatsApp Image)"

                    # WRITE CELLS
                    c1 = ws.cell(row=row_idx, column=col, value=f"CONTENT: {tag}")
                    c1.font = header_font; c1.fill = header_fill; c1.alignment = Alignment(horizontal='center')
                    
                    c2 = ws.cell(row=row_idx+1, column=col, value=f"FILE: {p_name} | DATE: {date_val}")
                    c2.font = label_font; c2.border = border
                    
                    c3 = ws.cell(row=row_idx+2, column=col, value=f"LOCATION: {loc_str}")
                    c3.font = label_font; c3.border = border
                    if lat: c3.hyperlink = f"https://www.google.com/maps?q={lat},{lon}"; c3.font = Font(color="0000FF", underline="single", size=9)

                    # INSERT IMAGE
                    try:
                        img_pil = Image.open(p_path)
                        # Fix Orientation if EXIF exists
                        try:
                            for orientation in ExifTags.TAGS.keys():
                                if ExifTags.TAGS[orientation]=='Orientation': break
                            exif=dict(img_pil._getexif().items())
                            if exif[orientation] == 3: img_pil=img_pil.rotate(180, expand=True)
                            elif exif[orientation] == 6: img_pil=img_pil.rotate(270, expand=True)
                            elif exif[orientation] == 8: img_pil=img_pil.rotate(90, expand=True)
                        except: pass

                        # Scale to fit approx 380px wide
                        target_w = 380
                        aspect = img_pil.height / img_pil.width
                        target_h = int(target_w * aspect)
                        img_pil = img_pil.resize((target_w, target_h), Image.Resampling.LANCZOS)
                        
                        img_byte_arr = io.BytesIO()
                        img_pil.save(img_byte_arr, format='JPEG')
                        img_byte_arr.seek(0)
                        
                        img_final = OpenpyxlImage(img_byte_arr)
                        # Anchor to cell
                        img_final.anchor = f"{chr(64+col)}{row_idx+3}"
                        ws.add_image(img_final)
                        
                        # Set row height to fit image (1 pixel = 0.75 points)
                        # Ensure we pick the max height if 2 photos are side-by-side
                        needed_h = (target_h * 0.75) + 10
                        if ws.row_dimensions[row_idx+3].height is None or needed_h > ws.row_dimensions[row_idx+3].height:
                            ws.row_dimensions[row_idx+3].height = needed_h

                    except Exception as e: 
                        print(f"Img Error: {e}")
                        pass

            row_idx += 5 # Move to next block: 3 text + 1 photo + 1 spacer
            ws.row_dimensions[row_idx-1].height = 15 # Spacer

        wb.save(excel_path)
        os.startfile(excel_path)
        messagebox.showinfo("Success", f"Task Complete! New sheet '{base_name}' added.")
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        transfer_btn.config(state="normal", text="Transfer Photos")

# ------------- GUI -------------
def select_folder():
    path = filedialog.askdirectory(); 
    if path: folder_var.set(path)

def select_excel():
    path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
    if path: excel_var.set(path)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("JKD Smart Photo Engine (AI Powered)")
    root.geometry("600x320")
    root.configure(bg="#f8fafc")
    
    style = ttk.Style(); style.theme_use('clam')
    
    tk.Label(root, text="JKD SITE PHOTO ANALYTICS (AI)", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#1e293b").pack(pady=15)
    
    # Folder
    tk.Label(root, text="Source Folder:", bg="#f8fafc").pack(anchor="w", padx=40)
    f_frame = tk.Frame(root, bg="#f8fafc")
    f_frame.pack(fill="x", padx=40)
    folder_var = tk.StringVar(value=r"D:\JKD_Folder\JKD-PROJECT SITE\JKD_SITE_PHOTOS")
    tk.Entry(f_frame, textvariable=folder_var, width=50).pack(side="left")
    tk.Button(f_frame, text="Browse", command=select_folder).pack(side="left", padx=5)

    # Excel
    tk.Label(root, text="Report File (History):", bg="#f8fafc").pack(anchor="w", padx=40, pady=(10,0))
    e_frame = tk.Frame(root, bg="#f8fafc")
    e_frame.pack(fill="x", padx=40)
    excel_var = tk.StringVar(value=r"D:\JKD_Folder\JKD-PROJECT SITE\JKD_Weekly_Progrees\Site_Photo_History.xlsx")
    tk.Entry(e_frame, textvariable=excel_var, width=50).pack(side="left")
    tk.Button(e_frame, text="Select", command=select_excel).pack(side="left", padx=5)

    transfer_btn = tk.Button(root, text="TRANSFER & AI ANALYZE", command=process_photos, bg="#3b82f6", fg="white", font=("Segoe UI", 12, "bold"), padx=30, pady=10, relief="flat")
    transfer_btn.pack(pady=25)

    if not AI_AVAILABLE:
        tk.Label(root, text="* AI Recognition initializing... please wait 1 min on first run.", font=("Arial", 8, "italic"), bg="#f8fafc", fg="#64748B").pack()

    root.mainloop()
