import os
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

excel_path = "test_full_logic.xlsx"
folder_path = r"D:\JKD_Folder\JKD-PROJECT SITE\JKD_SITE_PHOTOS\WPR_09-2-26\28-2-26"
photos = [f for f in os.listdir(folder_path) if f.lower().endswith('.jpeg')][:3]

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1, value="Photo Name")
ws.cell(row=1, column=2, value="Image (12cm x 18cm)")

row_idx = 2
for photo_name in photos:
    photo_path = os.path.join(folder_path, photo_name)
    img = OpenpyxlImage(photo_path)
    img.width = 454
    img.height = 680
    cell_ref = f"B{row_idx}"
    ws.row_dimensions[row_idx].height = 520
    ws.add_image(img, cell_ref)
    ws.cell(row=row_idx, column=1, value=photo_name)
    row_idx += 1

wb.save(excel_path)
print("Saved test_full_logic.xlsx")
