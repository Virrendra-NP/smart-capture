"""
===========================================================
MONTHLY CONSOLIDATED REPORT WITH DASHBOARD
===========================================================
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ============================================================
# SECTION 1: CHANGE DATES HERE - EDIT THIS DICTIONARY
# ============================================================

date_cols = {
    3: '01-Feb', 5: '02-Feb', 7: '03-Feb', 9: '04-Feb', 11: '05-Feb', 13: '06-Feb', 15: '07-Feb',
    17: '09-Feb', 19: '10-Feb', 21: '11-Feb', 23: '12-Feb', 25: '13-Feb', 27: '14-Feb',
    31: '17-Feb', 33: '18-Feb', 35: '19-Feb', 39: '20-Feb',
    41: '21-Feb', 43: '23-Feb', 45: '24-Feb', 47: '25-Feb', 49: '26-Feb', 51: '27-Feb', 53: '28-Feb'
}

# ============================================================
# SECTION 2: PROCESS DATA
# ============================================================

df = pd.read_excel(r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\JKD - DPR - 28.02.2026 (1) (1) (1).xlsx', sheet_name="Feb'26", header=None)

def get_weekly_data(start_row, end_row):
    data_rows = []
    for i in range(start_row, end_row):
        activity = df.iloc[i, 1]
        unit = df.iloc[i, 2]
        if pd.notna(activity) and activity not in ['Main Factory Building', 'Compound Wall'] and not isinstance(activity, float):
            p, a = 0, 0
            for col in date_cols.keys():
                p += float(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else 0
                a += float(df.iloc[i, col + 1]) if pd.notna(df.iloc[i, col + 1]) else 0
            if p > 0 or a > 0:
                data_rows.append({'Activity': activity, 'Unit': unit, 'Planned': p, 'Achieved': a, 'Variance': a - p})
    return pd.DataFrame(data_rows)

main_building = get_weekly_data(7, 26)
compound_wall = get_weekly_data(27, 40)

wb = Workbook()

# ============================================================
# SECTION 3: CREATE DASHBOARD
# ============================================================
ws_dash = wb.active
ws_dash.title = "DASHBOARD"
ws_dash.sheet_view.showGridLines = False

# Colors
heading_bg = "1F4E79"
kpi_bg_pct_good = "C6EFCE"
kpi_font_pct_good = "006100"
kpi_bg_pct_warn = "FFC7CE"
kpi_font_pct_warn = "9C0006"

# Fill sheet with white background
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
for r in range(1, 45):
    for c in range(1, 26):
        ws_dash.cell(row=r, column=c).fill = white_fill

# --- DASHBOARD HEADER ---
ws_dash['B2'] = "MONTHLY PROGRESS REPORT DASHBOARD - FEBRUARY 2026"
ws_dash['B2'].font = Font(bold=True, size=24, color='FFFFFF')
ws_dash['B2'].fill = PatternFill(start_color=heading_bg, end_color=heading_bg, fill_type='solid')
ws_dash.merge_cells('B2:Y4')
ws_dash['B2'].alignment = Alignment(horizontal='center', vertical='center')

ws_dash['B5'] = "Project: JKD 358 - PRODUCTION BLOCK"
ws_dash['B5'].font = Font(bold=True, size=14, color=heading_bg)
ws_dash.merge_cells('B5:Y5')
ws_dash['B5'].alignment = Alignment(horizontal='center', vertical='center')

# Calculations for KPIs
mb_planned = main_building['Planned'].sum()
mb_achieved = main_building['Achieved'].sum()
mb_pct = round(mb_achieved/mb_planned*100, 1) if mb_planned>0 else 0

cw_planned = compound_wall['Planned'].sum()
cw_achieved = compound_wall['Achieved'].sum()
cw_pct = round(cw_achieved/cw_planned*100, 1) if cw_planned>0 else 0

total_planned = mb_planned + cw_planned
total_achieved = mb_achieved + cw_achieved
total_pct = round(total_achieved/total_planned*100, 1) if total_planned>0 else 0

# Function to draw KPI Cards
def draw_kpi_card(ws, start_row, start_col, title, value, subtitle, bg_color="FFFFFF", font_color="000000"):
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+3)
    c = ws.cell(row=start_row, column=start_col, value=title)
    c.font = Font(bold=True, color="64748B", size=10)
    c.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+2, end_column=start_col+3)
    v = ws.cell(row=start_row+1, column=start_col, value=value)
    v.font = Font(bold=True, size=22, color=font_color)
    v.alignment = Alignment(horizontal='center', vertical='center')
    v.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    
    border = Border(left=Side(style='thin', color="CCCCCC"), right=Side(style='thin', color="CCCCCC"),
                    top=Side(style='thin', color="CCCCCC"), bottom=Side(style='thin', color="CCCCCC"))
    for r in range(start_row, start_row+4):
        for c_idx in range(start_col, start_col+4):
            ws.cell(row=r, column=c_idx).border = border

    ws.merge_cells(start_row=start_row+3, start_column=start_col, end_row=start_row+3, end_column=start_col+3)
    s = ws.cell(row=start_row+3, column=start_col, value=subtitle)
    s.font = Font(size=9, color="A0A0A0")
    s.alignment = Alignment(horizontal='center', vertical='center')

# Overall KPIs row 8
ws_dash['B8'] = "OVERALL PROGRESS SUMMARY"
ws_dash['B8'].font = Font(bold=True, size=16, color="333333")
ws_dash.merge_cells('B8:Y8')

draw_kpi_card(ws_dash, 10, 3, "TOTAL PLANNED", round(total_planned,1), "Units", "F8FAFC", "1F4E79")
draw_kpi_card(ws_dash, 10, 8, "TOTAL ACHIEVED", round(total_achieved,1), "Units", "F8FAFC", "1F4E79")
draw_kpi_card(ws_dash, 10, 13, "ACHIEVEMENT %", f"{total_pct}%", "Overall Progress", 
              kpi_bg_pct_good if total_pct >= 85 else kpi_bg_pct_warn, 
              kpi_font_pct_good if total_pct >= 85 else kpi_font_pct_warn)

# Details Section
ws_dash['B16'] = "MAIN FACTORY BUILDING"
ws_dash['B16'].font = Font(bold=True, size=14, color="4472C4")
ws_dash.merge_cells('B16:M16')

ws_dash['O16'] = "COMPOUND WALL"
ws_dash['O16'].font = Font(bold=True, size=14, color="4472C4")
ws_dash.merge_cells('O16:Y16')

draw_kpi_card(ws_dash, 18, 2, "PLANNED", round(mb_planned,1), "Units", "F8FAFC", "1F4E79")
draw_kpi_card(ws_dash, 18, 6, "ACHIEVED", round(mb_achieved,1), "Units", "F8FAFC", "1F4E79")
draw_kpi_card(ws_dash, 18, 10, "ACHIEVEMENT %", f"{mb_pct}%", "Main Building", 
              kpi_bg_pct_good if mb_pct >= 85 else kpi_bg_pct_warn, 
              kpi_font_pct_good if mb_pct >= 85 else kpi_font_pct_warn)

draw_kpi_card(ws_dash, 18, 15, "PLANNED", round(cw_planned,1), "Units", "F8FAFC", "1F4E79")
draw_kpi_card(ws_dash, 18, 19, "ACHIEVED", round(cw_achieved,1), "Units", "F8FAFC", "1F4E79")
draw_kpi_card(ws_dash, 18, 23, "ACHIEVEMENT %", f"{cw_pct}%", "Compound Wall", 
              kpi_bg_pct_good if cw_pct >= 85 else kpi_bg_pct_warn, 
              kpi_font_pct_good if cw_pct >= 85 else kpi_font_pct_warn)

# Set Column width to approx 5
for c in range(1, 30):
    ws_dash.column_dimensions[get_column_letter(c)].width = 4.5
ws_dash.column_dimensions['A'].width = 2


# ================= CHARTS DATA HIDDEN SHEET =================
ws_data = wb.create_sheet("Chart Data")
ws_data.sheet_state = 'hidden'

# Write MB data
ws_data['A1'] = "Activity (Unit)"
ws_data['B1'] = "Planned"
ws_data['C1'] = "Achieved"
ws_data['D1'] = "Variation"

c_row = 2
for idx, r in main_building.iterrows():
    # Adding unit to the activity label for the x-axis
    act_label = f"{r['Activity'][:25]}\n({r['Unit']})"
    ws_data.cell(row=c_row, column=1, value=act_label)
    ws_data.cell(row=c_row, column=2, value=r['Planned'])
    ws_data.cell(row=c_row, column=3, value=r['Achieved'])
    ws_data.cell(row=c_row, column=4, value=r['Variance'])
    c_row += 1
mb_end_row = c_row - 1

# Write CW data
ws_data['F1'] = "Activity (Unit)"
ws_data['G1'] = "Planned"
ws_data['H1'] = "Achieved"
ws_data['I1'] = "Variation"
c_row = 2
for idx, r in compound_wall.iterrows():
    act_label = f"{r['Activity'][:25]}\n({r['Unit']})"
    ws_data.cell(row=c_row, column=6, value=act_label)
    ws_data.cell(row=c_row, column=7, value=r['Planned'])
    ws_data.cell(row=c_row, column=8, value=r['Achieved'])
    ws_data.cell(row=c_row, column=9, value=r['Variance'])
    c_row += 1
cw_end_row = c_row - 1

# --- CREATE DASHBOARD CHARTS ---
def create_chart(title, min_col, max_col, min_row, max_row, cats_min_col):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Quantity"
    chart.grouping = "clustered"
    chart.overlap = 0 # Bars side by side, no overlap
    chart.gapWidth = 150 # Adjust gap for cleaner look
    
    data = Reference(ws_data, min_col=min_col, min_row=min_row, max_row=max_row, max_col=max_col)
    cats = Reference(ws_data, min_col=cats_min_col, min_row=min_row+1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 22 # Wider chart for labels
    chart.height = 12
    
    # Setup Colors & Labels
    colors = ["002060", "00B050", "C00000"] # Dark Blue, Green, Red
    for i, series in enumerate(chart.series):
        series.graphicalProperties.solidFill = colors[i % len(colors)]
        
        # Adding Data Labels (show quantities on top of bars)
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showCatName = False
        series.dLbls.position = "outEnd"
        
    return chart

# Add Chart 1
chart1 = create_chart("Main Building - Progress Detail", 2, 4, 1, mb_end_row, 1)
ws_dash.add_chart(chart1, "B24")

# Add Chart 2
chart2 = create_chart("Compound Wall - Progress Detail", 7, 9, 1, cw_end_row, 6)
ws_dash.add_chart(chart2, "O24")

# ============================================================
# CREATE INDIVIDUAL TABS WITH TABLES (From Original Script)
# ============================================================

def format_data_sheet(ws_name, data_df, title):
    ws = wb.create_sheet(ws_name)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = "JKD 358 - PRODUCTION BLOCK"
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    ws['A2'] = f"{title} - Monthly Progress: February 2026"
    ws['A2'].font = Font(size=12, italic=True)
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = title.upper()
    ws['A4'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.merge_cells('A4:G4')
    ws['A4'].alignment = Alignment(horizontal='center')
    
    headers = ['S.No', 'Activity', 'Unit', 'Planned', 'Achieved', 'Variance', 'Achievement %']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
    row = 6
    for idx, r in data_df.iterrows():
        ws.cell(row=row, column=1, value=idx + 1).border = border
        ws.cell(row=row, column=2, value=r['Activity']).border = border
        ws.cell(row=row, column=3, value=r['Unit']).border = border
        ws.cell(row=row, column=4, value=r['Planned']).border = border
        ws.cell(row=row, column=5, value=r['Achieved']).border = border
        var_cell = ws.cell(row=row, column=6, value=r['Variance'])
        var_cell.border = border
        if r['Variance'] < 0:
            var_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        else:
            var_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        pct_cell = ws.cell(row=row, column=7)
        pct_cell.value = f'=IF(D{row}>0,E{row}/D{row}*100,0)'
        pct_cell.border = border
        pct_cell.number_format = '0.0'
        row += 1

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 15

format_data_sheet("Main Building Data", main_building, "Main Factory Building")
format_data_sheet("Compound Wall Data", compound_wall, "Compound Wall")

# ============================================================
# OUTPUT FILENAME
# ============================================================

output_path = r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\Monthly_Report_Feb_2026_Dashboard.xlsx'
wb.save(output_path)
print(f"Excel Dashboard saved: {output_path}")

# Open the excel file automatically
try:
    os.startfile(output_path)
except Exception as e:
    print(f"Could not open file automatically: {e}")
