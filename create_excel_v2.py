import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.drawing.fill import ColorChoice

df = pd.read_excel(r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\JKD - DPR - 28.02.2026 (1) (1) (1).xlsx', sheet_name="Feb'26", header=None)

date_cols = {41: '21-Feb', 43: '23-Feb', 45: '24-Feb', 47: '25-Feb', 49: '26-Feb', 51: '27-Feb', 53: '28-Feb'}

def get_weekly_data(start_row, end_row):
    data_rows = []
    for i in range(start_row, end_row):
        activity = df.iloc[i, 1]
        unit = df.iloc[i, 2]
        if pd.notna(activity) and activity not in ['Main Factory Building', 'Compound Wall'] and not isinstance(activity, float):
            p, a = 0, 0
            for col in date_cols.keys():
                p += float(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else 0
                a += float(df.iloc[i, col + 1]) if pd.notna(df.iloc[i, col + 1]) else 0
            if p > 0 or a > 0:
                data_rows.append({'Activity': activity, 'Unit': unit, 'Planned': p, 'Achieved': a, 'Variance': a - p})
    return pd.DataFrame(data_rows)

main_building = get_weekly_data(7, 26)
compound_wall = get_weekly_data(27, 40)

wb = Workbook()

ws_mb = wb.active
ws_mb.title = "Main Building"

border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws_mb['A1'] = "JKD 358 - PRODUCTION BLOCK"
ws_mb['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_mb['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_mb.merge_cells('A1:G1')
ws_mb['A1'].alignment = Alignment(horizontal='center')
ws_mb.row_dimensions[1].height = 25

ws_mb['A2'] = "Main Factory Building - Weekly Progress: 21-Feb-2026 to 28-Feb-2026"
ws_mb['A2'].font = Font(size=12, italic=True)
ws_mb.merge_cells('A2:G2')
ws_mb['A2'].alignment = Alignment(horizontal='center')

ws_mb['A4'] = "MAIN FACTORY BUILDING"
ws_mb['A4'].font = Font(bold=True, size=14, color='FFFFFF')
ws_mb['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_mb.merge_cells('A4:G4')
ws_mb['A4'].alignment = Alignment(horizontal='center')

headers = ['S.No', 'Activity', 'Unit', 'Planned', 'Achieved', 'Variance', 'Achievement %']
for col_idx, h in enumerate(headers, 1):
    cell = ws_mb.cell(row=5, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

row = 6
mb_start = row
for idx, r in main_building.iterrows():
    ws_mb.cell(row=row, column=1, value=idx + 1).border = border
    ws_mb.cell(row=row, column=2, value=r['Activity']).border = border
    ws_mb.cell(row=row, column=3, value=r['Unit']).border = border
    ws_mb.cell(row=row, column=4, value=r['Planned']).border = border
    ws_mb.cell(row=row, column=5, value=r['Achieved']).border = border
    var_cell = ws_mb.cell(row=row, column=6, value=r['Variance'])
    var_cell.border = border
    if r['Variance'] < 0:
        var_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    else:
        var_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    pct_cell = ws_mb.cell(row=row, column=7)
    pct_cell.value = f'=IF(D{row}>0,E{row}/D{row}*100,0)'
    pct_cell.border = border
    pct_cell.number_format = '0.0'
    row += 1

mb_end = row - 1

ws_mb.column_dimensions['A'].width = 8
ws_mb.column_dimensions['B'].width = 35
ws_mb.column_dimensions['C'].width = 10
ws_mb.column_dimensions['D'].width = 14
ws_mb.column_dimensions['E'].width = 14
ws_mb.column_dimensions['F'].width = 14
ws_mb.column_dimensions['G'].width = 15

mb_chart_row = row + 2
ws_mb.cell(row=mb_chart_row, column=1, value="S.No")
ws_mb.cell(row=mb_chart_row, column=2, value="Activity")
ws_mb.cell(row=mb_chart_row, column=3, value="Planned")
ws_mb.cell(row=mb_chart_row, column=4, value="Achieved")
ws_mb.cell(row=mb_chart_row, column=5, value="Variance")
ws_mb.cell(row=mb_chart_row, column=6, value="Achievement %")

for col in range(1, 7):
    ws_mb.cell(row=mb_chart_row, column=col).font = Font(bold=True, color='FFFFFF')
    ws_mb.cell(row=mb_chart_row, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_mb.cell(row=mb_chart_row, column=col).alignment = Alignment(horizontal='center')
    ws_mb.cell(row=mb_chart_row, column=col).border = border

chart_data_start = mb_chart_row + 1
for idx, r in main_building.iterrows():
    ws_mb.cell(row=chart_data_start + idx, column=1, value=idx + 1)
    ws_mb.cell(row=chart_data_start + idx, column=2, value=r['Activity'][:25])
    ws_mb.cell(row=chart_data_start + idx, column=3, value=r['Planned'])
    ws_mb.cell(row=chart_data_start + idx, column=4, value=r['Achieved'])
    ws_mb.cell(row=chart_data_start + idx, column=5, value=r['Variance'])
    ws_mb.cell(row=chart_data_start + idx, column=6, value=round(r['Achieved']/r['Planned']*100, 1) if r['Planned']>0 else 0)
    for col in range(1, 7):
        ws_mb.cell(row=chart_data_start + idx, column=col).border = border

mb_activities_count = len(main_building)

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Main Factory Building - Planned vs Achieved"
chart1.y_axis.title = "Quantity"
chart1.x_axis.title = "Activity"
chart1.grouping = "clustered"
chart1.overlap = 10

data = Reference(ws_mb, min_col=3, min_row=mb_chart_row, max_row=mb_chart_row + mb_activities_count, max_col=5)
cats = Reference(ws_mb, min_col=2, min_row=mb_chart_row+1, max_row=mb_chart_row + mb_activities_count)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.width = 18
chart1.height = 12

s1 = chart1.series[0]
s1.graphicalProperties.solidFill = "5B9BD5"
s1.graphicalProperties.line.solidFill = "5B9BD5"

s2 = chart1.series[1]
s2.graphicalProperties.solidFill = "70AD47"
s2.graphicalProperties.line.solidFill = "70AD47"

s3 = chart1.series[2]
s3.graphicalProperties.solidFill = "C00000"
s3.graphicalProperties.line.solidFill = "C00000"

ws_mb.add_chart(chart1, f"A{mb_chart_row + mb_activities_count + 2}")

ws_cw = wb.create_sheet("Compound Wall")

ws_cw['A1'] = "JKD 358 - PRODUCTION BLOCK"
ws_cw['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_cw['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_cw.merge_cells('A1:G1')
ws_cw['A1'].alignment = Alignment(horizontal='center')
ws_cw.row_dimensions[1].height = 25

ws_cw['A2'] = "Compound Wall - Weekly Progress: 21-Feb-2026 to 28-Feb-2026"
ws_cw['A2'].font = Font(size=12, italic=True)
ws_cw.merge_cells('A2:G2')
ws_cw['A2'].alignment = Alignment(horizontal='center')

ws_cw['A4'] = "COMPOUND WALL"
ws_cw['A4'].font = Font(bold=True, size=14, color='FFFFFF')
ws_cw['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_cw.merge_cells('A4:G4')
ws_cw['A4'].alignment = Alignment(horizontal='center')

for col_idx, h in enumerate(headers, 1):
    cell = ws_cw.cell(row=5, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

row = 6
for idx, r in compound_wall.iterrows():
    ws_cw.cell(row=row, column=1, value=idx + 1).border = border
    ws_cw.cell(row=row, column=2, value=r['Activity']).border = border
    ws_cw.cell(row=row, column=3, value=r['Unit']).border = border
    ws_cw.cell(row=row, column=4, value=r['Planned']).border = border
    ws_cw.cell(row=row, column=5, value=r['Achieved']).border = border
    var_cell = ws_cw.cell(row=row, column=6, value=r['Variance'])
    var_cell.border = border
    if r['Variance'] < 0:
        var_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    else:
        var_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    pct_cell = ws_cw.cell(row=row, column=7)
    pct_cell.value = f'=IF(D{row}>0,E{row}/D{row}*100,0)'
    pct_cell.border = border
    pct_cell.number_format = '0.0'
    row += 1

cw_end = row - 1

ws_cw.column_dimensions['A'].width = 8
ws_cw.column_dimensions['B'].width = 35
ws_cw.column_dimensions['C'].width = 10
ws_cw.column_dimensions['D'].width = 14
ws_cw.column_dimensions['E'].width = 14
ws_cw.column_dimensions['F'].width = 14
ws_cw.column_dimensions['G'].width = 15

cw_chart_row = row + 2
ws_cw.cell(row=cw_chart_row, column=1, value="S.No")
ws_cw.cell(row=cw_chart_row, column=2, value="Activity")
ws_cw.cell(row=cw_chart_row, column=3, value="Planned")
ws_cw.cell(row=cw_chart_row, column=4, value="Achieved")
ws_cw.cell(row=cw_chart_row, column=5, value="Variance")
ws_cw.cell(row=cw_chart_row, column=6, value="Achievement %")

for col in range(1, 7):
    ws_cw.cell(row=cw_chart_row, column=col).font = Font(bold=True, color='FFFFFF')
    ws_cw.cell(row=cw_chart_row, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_cw.cell(row=cw_chart_row, column=col).alignment = Alignment(horizontal='center')
    ws_cw.cell(row=cw_chart_row, column=col).border = border

chart_cw_start = cw_chart_row + 1
for idx, r in compound_wall.iterrows():
    ws_cw.cell(row=chart_cw_start + idx, column=1, value=idx + 1)
    ws_cw.cell(row=chart_cw_start + idx, column=2, value=r['Activity'][:25])
    ws_cw.cell(row=chart_cw_start + idx, column=3, value=r['Planned'])
    ws_cw.cell(row=chart_cw_start + idx, column=4, value=r['Achieved'])
    ws_cw.cell(row=chart_cw_start + idx, column=5, value=r['Variance'])
    ws_cw.cell(row=chart_cw_start + idx, column=6, value=round(r['Achieved']/r['Planned']*100, 1) if r['Planned']>0 else 0)
    for col in range(1, 7):
        ws_cw.cell(row=chart_cw_start + idx, column=col).border = border

cw_activities_count = len(compound_wall)

chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.title = "Compound Wall - Planned vs Achieved"
chart2.y_axis.title = "Quantity"
chart2.x_axis.title = "Activity"
chart2.grouping = "clustered"
chart2.overlap = 10

data2 = Reference(ws_cw, min_col=3, min_row=cw_chart_row, max_row=cw_chart_row + cw_activities_count, max_col=5)
cats2 = Reference(ws_cw, min_col=2, min_row=cw_chart_row+1, max_row=cw_chart_row + cw_activities_count)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 18
chart2.height = 12

s1 = chart2.series[0]
s1.graphicalProperties.solidFill = "5B9BD5"
s1.graphicalProperties.line.solidFill = "5B9BD5"

s2 = chart2.series[1]
s2.graphicalProperties.solidFill = "70AD47"
s2.graphicalProperties.line.solidFill = "70AD47"

s3 = chart2.series[2]
s3.graphicalProperties.solidFill = "C00000"
s3.graphicalProperties.line.solidFill = "C00000"

ws_cw.add_chart(chart2, f"A{cw_chart_row + cw_activities_count + 2}")

ws_kpi = wb.create_sheet("KPI Summary")

ws_kpi['A1'] = "KPI DASHBOARD - JKD 358 Production Block"
ws_kpi['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_kpi['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_kpi.merge_cells('A1:F1')
ws_kpi['A1'].alignment = Alignment(horizontal='center')
ws_kpi.row_dimensions[1].height = 25

ws_kpi['A2'] = "Weekly Progress: 21-Feb-2026 to 28-Feb-2026"
ws_kpi['A2'].font = Font(size=11, italic=True)
ws_kpi.merge_cells('A2:F2')
ws_kpi['A2'].alignment = Alignment(horizontal='center')

mb_planned = main_building['Planned'].sum()
mb_achieved = main_building['Achieved'].sum()
mb_pct = round(mb_achieved/mb_planned*100, 1) if mb_planned>0 else 0
cw_planned = compound_wall['Planned'].sum()
cw_achieved = compound_wall['Achieved'].sum()
cw_pct = round(cw_achieved/cw_planned*100, 1) if cw_planned>0 else 0
total_planned = mb_planned + cw_planned
total_achieved = mb_achieved + cw_achieved
total_pct = round(total_achieved/total_planned*100, 1) if total_planned>0 else 0

ws_kpi['A4'] = "MAIN FACTORY BUILDING - KPI"
ws_kpi['A4'].font = Font(bold=True, size=12, color='FFFFFF')
ws_kpi['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_kpi.merge_cells('A4:C4')
ws_kpi['A4'].alignment = Alignment(horizontal='center')

kpi_headers = ['Metric', 'Value']
for col, h in enumerate(kpi_headers, 1):
    ws_kpi.cell(row=5, column=col, value=h).font = Font(bold=True)
    ws_kpi.cell(row=5, column=col).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    ws_kpi.cell(row=5, column=col).border = border

mb_kpis = [
    ['Total Planned', round(mb_planned, 2)],
    ['Total Achieved', round(mb_achieved, 2)],
    ['Achievement %', f"{mb_pct}%"],
]

for idx, (label, val) in enumerate(mb_kpis, 6):
    ws_kpi.cell(row=idx, column=1, value=label).border = border
    ws_kpi.cell(row=idx, column=2, value=val).border = border
    if 'Achievement' in label:
        ws_kpi.cell(row=idx, column=2).font = Font(bold=True, size=12, color='1F4E79')

ws_kpi['A11'] = "COMPOUND WALL - KPI"
ws_kpi['A11'].font = Font(bold=True, size=12, color='FFFFFF')
ws_kpi['A11'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_kpi.merge_cells('A11:C11')
ws_kpi['A11'].alignment = Alignment(horizontal='center')

for col, h in enumerate(kpi_headers, 1):
    ws_kpi.cell(row=12, column=col, value=h).font = Font(bold=True)
    ws_kpi.cell(row=12, column=col).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    ws_kpi.cell(row=12, column=col).border = border

cw_kpis = [
    ['Total Planned', round(cw_planned, 2)],
    ['Total Achieved', round(cw_achieved, 2)],
    ['Achievement %', f"{cw_pct}%"],
]

for idx, (label, val) in enumerate(cw_kpis, 13):
    ws_kpi.cell(row=idx, column=1, value=label).border = border
    ws_kpi.cell(row=idx, column=2, value=val).border = border
    if 'Achievement' in label:
        ws_kpi.cell(row=idx, column=2).font = Font(bold=True, size=12, color='1F4E79')

ws_kpi['A18'] = "OVERALL - KPI"
ws_kpi['A18'].font = Font(bold=True, size=12, color='FFFFFF')
ws_kpi['A18'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_kpi.merge_cells('A18:C18')
ws_kpi['A18'].alignment = Alignment(horizontal='center')

for col, h in enumerate(kpi_headers, 1):
    ws_kpi.cell(row=19, column=col, value=h).font = Font(bold=True)
    ws_kpi.cell(row=19, column=col).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    ws_kpi.cell(row=19, column=col).border = border

overall_kpis = [
    ['Total Planned', round(total_planned, 2)],
    ['Total Achieved', round(total_achieved, 2)],
    ['Achievement %', f"{total_pct}%"],
]

for idx, (label, val) in enumerate(overall_kpis, 20):
    ws_kpi.cell(row=idx, column=1, value=label).border = border
    ws_kpi.cell(row=idx, column=2, value=val).border = border
    if 'Achievement' in label:
        ws_kpi.cell(row=idx, column=2).font = Font(bold=True, size=14, color='1F4E79')

ws_kpi.column_dimensions['A'].width = 20
ws_kpi.column_dimensions['B'].width = 18

chart_row_kpi = 4
ws_kpi.cell(row=chart_row_kpi, column=5, value="Section")
ws_kpi.cell(row=chart_row_kpi, column=6, value="Achievement %")
ws_kpi.cell(row=chart_row_kpi+1, column=5, value="Main Building")
ws_kpi.cell(row=chart_row_kpi+1, column=6, value=mb_pct)
ws_kpi.cell(row=chart_row_kpi+2, column=5, value="Compound Wall")
ws_kpi.cell(row=chart_row_kpi+2, column=6, value=cw_pct)

for col in range(5, 7):
    ws_kpi.cell(row=chart_row_kpi, column=col).font = Font(bold=True, color='FFFFFF')
    ws_kpi.cell(row=chart_row_kpi, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_kpi.cell(row=chart_row_kpi, column=col).alignment = Alignment(horizontal='center')
    ws_kpi.cell(row=chart_row_kpi, column=col).border = border

chart_kpi = BarChart()
chart_kpi.type = "col"
chart_kpi.title = "Achievement % by Section"
chart_kpi.y_axis.title = "Percentage"
chart_kpi.y_axis.scaling.min = 0
chart_kpi.y_axis.scaling.max = 120

data_kpi = Reference(ws_kpi, min_col=6, min_row=chart_row_kpi, max_row=chart_row_kpi+2)
cats_kpi = Reference(ws_kpi, min_col=5, min_row=chart_row_kpi+1, max_row=chart_row_kpi+2)
chart_kpi.add_data(data_kpi, titles_from_data=True)
chart_kpi.set_categories(cats_kpi)
chart_kpi.width = 12
chart_kpi.height = 8

ws_kpi.add_chart(chart_kpi, "E4")

output_path = r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\Weekly_Progress_Report_21-28_Feb_2026.xlsx'
wb.save(output_path)
print(f"Excel saved: {output_path}")
print(f"\n=== SUMMARY ===")
print(f"Main Building - Planned: {mb_planned:.2f}, Achieved: {mb_achieved:.2f}, %: {mb_pct}%")
print(f"Compound Wall - Planned: {cw_planned:.2f}, Achieved: {cw_achieved:.2f}, %: {cw_pct}%")
print(f"Total - Planned: {total_planned:.2f}, Achieved: {total_achieved:.2f}, %: {total_pct}%")
