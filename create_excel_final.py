import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

df = pd.read_excel(r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\JKD - DPR - 28.02.2026 (1) (1) (1).xlsx', sheet_name="Feb'26", header=None)

date_cols = {41: '21-Feb', 43: '23-Feb', 45: '24-Feb', 47: '25-Feb', 49: '26-Feb', 51: '27-Feb', 53: '28-Feb'}

def get_weekly_data(start_row, end_row):
    data_rows = []
    for i in range(start_row, end_row):
        activity = df.iloc[i, 1]
        unit = df.iloc[i, 2]
        if pd.notna(activity) and activity not in ['Main Factory Building', 'Compound Wall'] and not isinstance(activity, float):
            p, a = 0, 0
            for col in date_cols.keys():
                p += float(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else 0
                a += float(df.iloc[i, col + 1]) if pd.notna(df.iloc[i, col + 1]) else 0
            if p > 0 or a > 0:
                data_rows.append({'Activity': activity, 'Unit': unit, 'Planned': p, 'Achieved': a})
    return pd.DataFrame(data_rows)

main_building = get_weekly_data(7, 26)
compound_wall = get_weekly_data(27, 40)

wb = Workbook()
ws_data = wb.active
ws_data.title = "Weekly Progress"

ws_data['A1'] = "JKD 358 - PRODUCTION BLOCK"
ws_data['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_data['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_data.merge_cells('A1:F1')
ws_data['A1'].alignment = Alignment(horizontal='center')
ws_data.row_dimensions[1].height = 25

ws_data['A2'] = "Daily Progress Report: 21-Feb-2026 to 28-Feb-2026"
ws_data['A2'].font = Font(size=11, italic=True)
ws_data.merge_cells('A2:F2')
ws_data['A2'].alignment = Alignment(horizontal='center')

border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws_data['A4'] = "MAIN FACTORY BUILDING"
ws_data['A4'].font = Font(bold=True, size=13, color='FFFFFF')
ws_data['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_data.merge_cells('A4:F4')
ws_data['A4'].alignment = Alignment(horizontal='center')

headers = ['Activity', 'Unit', 'Planned', 'Achieved', 'Variance', 'Achievement %']
for col_idx, h in enumerate(headers, 1):
    cell = ws_data.cell(row=5, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

row = 6
mb_start = row
for idx, r in main_building.iterrows():
    ws_data.cell(row=row, column=1, value=r['Activity']).border = border
    ws_data.cell(row=row, column=2, value=r['Unit']).border = border
    ws_data.cell(row=row, column=3, value=r['Planned']).border = border
    ws_data.cell(row=row, column=4, value=r['Achieved']).border = border
    var_cell = ws_data.cell(row=row, column=5)
    var_cell.value = f'=D{row}-C{row}'
    var_cell.border = border
    pct_cell = ws_data.cell(row=row, column=6)
    pct_cell.value = f'=IF(C{row}>0,D{row}/C{row}*100,0)'
    pct_cell.border = border
    row += 1

mb_end = row - 1

ws_data.cell(row=row, column=1, value="TOTAL - MAIN BUILDING").font = Font(bold=True, size=11)
ws_data.cell(row=row, column=3, value=f'=SUM(C{mb_start}:C{mb_end})')
ws_data.cell(row=row, column=3).font = Font(bold=True)
ws_data.cell(row=row, column=4, value=f'=SUM(D{mb_start}:D{mb_end})')
ws_data.cell(row=row, column=4).font = Font(bold=True)
ws_data.cell(row=row, column=5, value=f'=D{row}-C{row}')
ws_data.cell(row=row, column=5).font = Font(bold=True)
ws_data.cell(row=row, column=6, value=f'=IF(C{row}>0,D{row}/C{row}*100,0)')
ws_data.cell(row=row, column=6).font = Font(bold=True, size=11)
ws_data.cell(row=row, column=6).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')

for col in range(1, 7):
    ws_data.cell(row=row, column=col).border = border

row += 2

ws_data.cell(row=row, column=1, value="COMPOUND WALL")
ws_data.cell(row=row, column=1).font = Font(bold=True, size=13, color='FFFFFF')
ws_data.cell(row=row, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_data.merge_cells(f'A{row}:F{row}')
ws_data.cell(row=row, column=1).alignment = Alignment(horizontal='center')
row += 1

for col_idx, h in enumerate(headers, 1):
    cell = ws_data.cell(row=row, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

row += 1
cw_start = row

for idx, r in compound_wall.iterrows():
    ws_data.cell(row=row, column=1, value=r['Activity']).border = border
    ws_data.cell(row=row, column=2, value=r['Unit']).border = border
    ws_data.cell(row=row, column=3, value=r['Planned']).border = border
    ws_data.cell(row=row, column=4, value=r['Achieved']).border = border
    var_cell = ws_data.cell(row=row, column=5)
    var_cell.value = f'=D{row}-C{row}'
    var_cell.border = border
    pct_cell = ws_data.cell(row=row, column=6)
    pct_cell.value = f'=IF(C{row}>0,D{row}/C{row}*100,0)'
    pct_cell.border = border
    row += 1

cw_end = row - 1

ws_data.cell(row=row, column=1, value="TOTAL - COMPOUND WALL").font = Font(bold=True, size=11)
ws_data.cell(row=row, column=3, value=f'=SUM(C{cw_start}:C{cw_end})')
ws_data.cell(row=row, column=3).font = Font(bold=True)
ws_data.cell(row=row, column=4, value=f'=SUM(D{cw_start}:D{cw_end})')
ws_data.cell(row=row, column=4).font = Font(bold=True)
ws_data.cell(row=row, column=5, value=f'=D{row}-C{row}')
ws_data.cell(row=row, column=5).font = Font(bold=True)
ws_data.cell(row=row, column=6, value=f'=IF(C{row}>0,D{row}/C{row}*100,0)')
ws_data.cell(row=row, column=6).font = Font(bold=True, size=11)
ws_data.cell(row=row, column=6).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')

for col in range(1, 7):
    ws_data.cell(row=row, column=col).border = border

grand_total_row = row + 2
ws_data.cell(row=grand_total_row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=12, color='FFFFFF')
ws_data.cell(row=grand_total_row, column=1).fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_data.merge_cells(f'A{grand_total_row}:B{grand_total_row}')
ws_data.cell(row=grand_total_row, column=3, value=f'=C{row-1}+C{mb_end+1}')
ws_data.cell(row=grand_total_row, column=4, value=f'=D{row-1}+D{mb_end+1}')
ws_data.cell(row=grand_total_row, column=5, value=f'=E{row-1}+E{mb_end+1}')
ws_data.cell(row=grand_total_row, column=6, value=f'=IF(C{grand_total_row}>0,D{grand_total_row}/C{grand_total_row}*100,0)')
for col in range(3, 7):
    ws_data.cell(row=grand_total_row, column=col).font = Font(bold=True, size=12, color='FFFFFF')
    ws_data.cell(row=grand_total_row, column=col).fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    ws_data.cell(row=grand_total_row, column=col).border = border

ws_data.column_dimensions['A'].width = 35
ws_data.column_dimensions['B'].width = 10
ws_data.column_dimensions['C'].width = 12
ws_data.column_dimensions['D'].width = 12
ws_data.column_dimensions['E'].width = 12
ws_data.column_dimensions['F'].width = 15

mb_planned = main_building['Planned'].sum()
mb_achieved = main_building['Achieved'].sum()
cw_planned = compound_wall['Planned'].sum()
cw_achieved = compound_wall['Achieved'].sum()
total_planned = mb_planned + cw_planned
total_achieved = mb_achieved + cw_achieved

ws_chart = wb.create_sheet("KPI Dashboard")

ws_chart['A1'] = "KPI DASHBOARD - JKD 358 Production Block"
ws_chart['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_chart['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_chart.merge_cells('A1:F1')
ws_chart['A1'].alignment = Alignment(horizontal='center')
ws_chart.row_dimensions[1].height = 25

ws_chart['A2'] = "Weekly Progress: 21-Feb-2026 to 28-Feb-2026"
ws_chart['A2'].font = Font(size=11, italic=True)
ws_chart.merge_cells('A2:F2')
ws_chart['A2'].alignment = Alignment(horizontal='center')

chart_row = 4
ws_chart.cell(row=chart_row, column=1, value="Section")
ws_chart.cell(row=chart_row, column=2, value="Planned")
ws_chart.cell(row=chart_row, column=3, value="Achieved")
ws_chart.cell(row=chart_row, column=4, value="Achievement %")

ws_chart.cell(row=chart_row+1, column=1, value="Main Factory Building")
ws_chart.cell(row=chart_row+1, column=2, value=mb_planned)
ws_chart.cell(row=chart_row+1, column=3, value=mb_achieved)
ws_chart.cell(row=chart_row+1, column=4, value=round(mb_achieved/mb_planned*100, 1) if mb_planned>0 else 0)

ws_chart.cell(row=chart_row+2, column=1, value="Compound Wall")
ws_chart.cell(row=chart_row+2, column=2, value=cw_planned)
ws_chart.cell(row=chart_row+2, column=3, value=cw_achieved)
ws_chart.cell(row=chart_row+2, column=4, value=round(cw_achieved/cw_planned*100, 1) if cw_planned>0 else 0)

ws_chart.cell(row=chart_row+3, column=1, value="TOTAL")
ws_chart.cell(row=chart_row+3, column=2, value=total_planned)
ws_chart.cell(row=chart_row+3, column=3, value=total_achieved)
ws_chart.cell(row=chart_row+3, column=4, value=round(total_achieved/total_planned*100, 1) if total_planned>0 else 0)

for col in range(1, 5):
    ws_chart.cell(row=chart_row, column=col).font = Font(bold=True, color='FFFFFF')
    ws_chart.cell(row=chart_row, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_chart.cell(row=chart_row, column=col).alignment = Alignment(horizontal='center')
    ws_chart.cell(row=chart_row, column=col).border = border
    
    ws_chart.cell(row=chart_row+3, column=col).font = Font(bold=True, size=11)
    ws_chart.cell(row=chart_row+3, column=col).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    ws_chart.cell(row=chart_row+3, column=col).border = border

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Planned vs Achieved - By Section"
chart1.y_axis.title = "Quantity (Units)"
chart1.x_axis.title = ""
chart1.legend.position = 'r'

data = Reference(ws_chart, min_col=2, min_row=chart_row, max_row=chart_row+3, max_col=3)
cats = Reference(ws_chart, min_col=1, min_row=chart_row+1, max_row=chart_row+3)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.width = 14
chart1.height = 10

ws_chart.add_chart(chart1, "A10")

pie_row = 10
ws_chart.cell(row=pie_row, column=5, value="Progress Status")
ws_chart.cell(row=pie_row, column=6, value="Units")
ws_chart.cell(row=pie_row+1, column=5, value="Achieved")
ws_chart.cell(row=pie_row+1, column=6, value=total_achieved)
ws_chart.cell(row=pie_row+2, column=5, value="Pending")
ws_chart.cell(row=pie_row+2, column=6, value=total_planned - total_achieved)

for col in range(5, 7):
    ws_chart.cell(row=pie_row, column=col).font = Font(bold=True, color='FFFFFF')
    ws_chart.cell(row=pie_row, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_chart.cell(row=pie_row, column=col).alignment = Alignment(horizontal='center')
    ws_chart.cell(row=pie_row, column=col).border = border
    ws_chart.cell(row=pie_row+1, column=col).border = border
    ws_chart.cell(row=pie_row+2, column=col).border = border

pie = PieChart()
pie.title = "Overall Progress"
labels = Reference(ws_chart, min_col=5, min_row=pie_row+1, max_row=pie_row+2)
pie_data = Reference(ws_chart, min_col=6, min_row=pie_row, max_row=pie_row+2)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 11
pie.height = 9

ws_chart.add_chart(pie, "E10")

act_chart_row = 25
ws_chart.cell(row=act_chart_row, column=1, value="Activity")
ws_chart.cell(row=act_chart_row, column=2, value="Planned")
ws_chart.cell(row=act_chart_row, column=3, value="Achieved")
ws_chart.cell(row=act_chart_row, column=4, value="Achievement %")

all_activities = pd.concat([main_building, compound_wall]).reset_index(drop=True)

for idx, (_, r) in enumerate(all_activities.iterrows(), act_chart_row+1):
    ws_chart.cell(row=idx, column=1, value=r['Activity'][:28])
    ws_chart.cell(row=idx, column=2, value=r['Planned'])
    ws_chart.cell(row=idx, column=3, value=r['Achieved'])
    ws_chart.cell(row=idx, column=4, value=round(r['Achieved']/r['Planned']*100, 1) if r['Planned']>0 else 0)

for col in range(1, 5):
    ws_chart.cell(row=act_chart_row, column=col).font = Font(bold=True, color='FFFFFF')
    ws_chart.cell(row=act_chart_row, column=col).fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    ws_chart.cell(row=act_chart_row, column=col).alignment = Alignment(horizontal='center')
    ws_chart.cell(row=act_chart_row, column=col).border = border

chart3 = BarChart()
chart3.type = "bar"
chart3.title = "Achievement % - By Activity"
chart3.y_axis.title = ""
chart3.x_axis.title = "Achievement %"

data = Reference(ws_chart, min_col=4, min_row=act_chart_row, max_row=act_chart_row + len(all_activities))
cats = Reference(ws_chart, min_col=1, min_row=act_chart_row+1, max_row=act_chart_row + len(all_activities))
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
chart3.width = 15
chart3.height = max(10, len(all_activities) * 0.5)

ws_chart.add_chart(chart3, "A35")

ws_chart.column_dimensions['A'].width = 30
ws_chart.column_dimensions['B'].width = 14
ws_chart.column_dimensions['C'].width = 14
ws_chart.column_dimensions['D'].width = 16
ws_chart.column_dimensions['E'].width = 18
ws_chart.column_dimensions['F'].width = 12

output_path = r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\Weekly_Progress_Report_21-28_Feb_2026.xlsx'
wb.save(output_path)
print(f"Excel saved: {output_path}")
print(f"\n=== SUMMARY ===")
print(f"Main Building - Planned: {mb_planned:.2f}, Achieved: {mb_achieved:.2f}, %: {round(mb_achieved/mb_planned*100, 1)}%")
print(f"Compound Wall - Planned: {cw_planned:.2f}, Achieved: {cw_achieved:.2f}, %: {round(cw_achieved/cw_planned*100, 1)}%")
print(f"Total - Planned: {total_planned:.2f}, Achieved: {total_achieved:.2f}, %: {round(total_achieved/total_planned*100, 1)}%")
