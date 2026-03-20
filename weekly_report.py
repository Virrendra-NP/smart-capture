"""
===========================================================
MONTHLY/WEEKLY REPORT GENERATOR - COMBINED CHARTS
===========================================================

TO CHANGE THE MONTH/PERIOD, EDIT THESE LINES ONLY:
----------------------------------------------------

LINE 14: date_cols - Add all dates and their column numbers
LINE 185: output_path - Change filename

===========================================================
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# ============================================================
# SECTION 1: CHANGE DATES HERE
# ============================================================

# For WEEKLY 21-28 Feb:
# date_cols = {41: '21-Feb', 43: '23-Feb', 45: '24-Feb', 47: '25-Feb', 49: '26-Feb', 51: '27-Feb', 53: '28-Feb'}

# For MONTHLY February (all dates):
date_cols = {
    3: '01-Feb', 5: '02-Feb', 7: '03-Feb', 9: '04-Feb', 11: '05-Feb', 13: '06-Feb', 15: '07-Feb',
    17: '09-Feb', 19: '10-Feb', 21: '11-Feb', 23: '12-Feb', 25: '13-Feb', 27: '14-Feb',
    31: '17-Feb', 33: '18-Feb', 35: '19-Feb', 39: '20-Feb',
    41: '21-Feb', 43: '23-Feb', 45: '24-Feb', 47: '25-Feb', 49: '26-Feb', 51: '27-Feb', 53: '28-Feb'
}

# ============================================================

df = pd.read_excel(r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\JKD - DPR - 28.02.2026 (1) (1) (1).xlsx', sheet_name="Feb'26", header=None)

def get_data(start_row, end_row):
    data_rows = []
    for i in range(start_row, end_row):
        activity = df.iloc[i, 1]
        unit = df.iloc[i, 2]
        if pd.notna(activity) and activity not in ['Main Factory Building', 'Compound Wall'] and not isinstance(activity, float):
            p, a = 0, 0
            for col in date_cols.keys():
                p += float(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else 0
                a += float(df.iloc[i, col + 1]) if pd.notna(df.iloc[i, col + 1]) else 0
            if p > 0 or a > 0:
                data_rows.append({'Activity': activity, 'Unit': unit, 'Planned': round(p,2), 'Achieved': round(a,2)})
    return pd.DataFrame(data_rows)

main_building = get_data(7, 26)
compound_wall = get_data(27, 40)

wb = Workbook()

# ============================================================
# MAIN BUILDING SHEET
# ============================================================

ws_mb = wb.active
ws_mb.title = "Main Building"

border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws_mb['A1'] = "JKD 358 - PRODUCTION BLOCK"
ws_mb['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_mb['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_mb.merge_cells('A1:G1')
ws_mb['A1'].alignment = Alignment(horizontal='center')
ws_mb.row_dimensions[1].height = 25

ws_mb['A2'] = "Main Factory Building - Monthly Progress: February 2026"
ws_mb['A2'].font = Font(size=11, italic=True)
ws_mb.merge_cells('A2:G2')
ws_mb['A2'].alignment = Alignment(horizontal='center')

ws_mb['A4'] = "MAIN FACTORY BUILDING"
ws_mb['A4'].font = Font(bold=True, size=14, color='FFFFFF')
ws_mb['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_mb.merge_cells('A4:G4')
ws_mb['A4'].alignment = Alignment(horizontal='center')

headers = ['S.No', 'Activity', 'Unit', 'Planned', 'Achieved', 'Variance', 'Achievement %']
for col_idx, h in enumerate(headers, 1):
    cell = ws_mb.cell(row=5, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

row = 6
for idx, r in main_building.iterrows():
    ws_mb.cell(row=row, column=1, value=idx + 1).border = border
    ws_mb.cell(row=row, column=2, value=r['Activity']).border = border
    ws_mb.cell(row=row, column=3, value=r['Unit']).border = border
    ws_mb.cell(row=row, column=4, value=r['Planned']).border = border
    ws_mb.cell(row=row, column=5, value=r['Achieved']).border = border
    
    var_cell = ws_mb.cell(row=row, column=6)
    var_cell.value = f'=E{row}-D{row}'
    var_cell.border = border
    
    pct_cell = ws_mb.cell(row=row, column=7)
    pct_cell.value = f'=IF(D{row}>0,E{row}/D{row}*100,0)'
    pct_cell.border = border
    pct_cell.number_format = '0.0'
    
    row += 1

mb_end = row - 1

ws_mb.column_dimensions['A'].width = 6
ws_mb.column_dimensions['B'].width = 32
ws_mb.column_dimensions['C'].width = 8
ws_mb.column_dimensions['D'].width = 12
ws_mb.column_dimensions['E'].width = 12
ws_mb.column_dimensions['F'].width = 12
ws_mb.column_dimensions['G'].width = 14

# COMBINED CHART - All 3 in one
mb_chart_start = row + 2
ws_mb.cell(row=mb_chart_start, column=1, value="Activity")
ws_mb.cell(row=mb_chart_start, column=2, value="Planned")
ws_mb.cell(row=mb_chart_start, column=3, value="Achieved")
ws_mb.cell(row=mb_chart_start, column=4, value="Variance")

for col in range(1, 5):
    ws_mb.cell(row=mb_chart_start, column=col).font = Font(bold=True, color='FFFFFF')
    ws_mb.cell(row=mb_chart_start, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_mb.cell(row=mb_chart_start, column=col).alignment = Alignment(horizontal='center')

for idx, r in main_building.iterrows():
    variance = r['Achieved'] - r['Planned']
    ws_mb.cell(row=mb_chart_start + idx + 1, column=1, value=r['Activity'][:22])
    ws_mb.cell(row=mb_chart_start + idx + 1, column=2, value=r['Planned'])
    ws_mb.cell(row=mb_chart_start + idx + 1, column=3, value=r['Achieved'])
    ws_mb.cell(row=mb_chart_start + idx + 1, column=4, value=variance)

chart_mb = BarChart()
chart_mb.type = "col"
chart_mb.style = 10
chart_mb.title = "Main Factory Building - Planned vs Achieved vs Variance"
chart_mb.y_axis.title = "Quantity"
chart_mb.x_axis.title = ""
chart_mb.grouping = "clustered"
chart_mb.overlap = -30

data_mb = Reference(ws_mb, min_col=2, min_row=mb_chart_start, max_row=mb_chart_start + len(main_building), max_col=4)
cats_mb = Reference(ws_mb, min_col=1, min_row=mb_chart_start+1, max_row=mb_chart_start + len(main_building))
chart_mb.add_data(data_mb, titles_from_data=True)
chart_mb.set_categories(cats_mb)
chart_mb.width = 28
chart_mb.height = 16

chart_mb.series[0].graphicalProperties.solidFill = "5B9BD5"
chart_mb.series[1].graphicalProperties.solidFill = "70AD47"
chart_mb.series[2].graphicalProperties.solidFill = "C00000"

chart_mb.dataLabels = DataLabelList()
chart_mb.dataLabels.showVal = True

ws_mb.add_chart(chart_mb, f"A{mb_chart_start + len(main_building) + 2}")

# ============================================================
# COMPOUND WALL SHEET
# ============================================================

ws_cw = wb.create_sheet("Compound Wall")

ws_cw['A1'] = "JKD 358 - PRODUCTION BLOCK"
ws_cw['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_cw['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_cw.merge_cells('A1:G1')
ws_cw['A1'].alignment = Alignment(horizontal='center')
ws_cw.row_dimensions[1].height = 25

ws_cw['A2'] = "Compound Wall - Monthly Progress: February 2026"
ws_cw['A2'].font = Font(size=11, italic=True)
ws_cw.merge_cells('A2:G2')
ws_cw['A2'].alignment = Alignment(horizontal='center')

ws_cw['A4'] = "COMPOUND WALL"
ws_cw['A4'].font = Font(bold=True, size=14, color='FFFFFF')
ws_cw['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_cw.merge_cells('A4:G4')
ws_cw['A4'].alignment = Alignment(horizontal='center')

for col_idx, h in enumerate(headers, 1):
    cell = ws_cw.cell(row=5, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

row = 6
for idx, r in compound_wall.iterrows():
    ws_cw.cell(row=row, column=1, value=idx + 1).border = border
    ws_cw.cell(row=row, column=2, value=r['Activity']).border = border
    ws_cw.cell(row=row, column=3, value=r['Unit']).border = border
    ws_cw.cell(row=row, column=4, value=r['Planned']).border = border
    ws_cw.cell(row=row, column=5, value=r['Achieved']).border = border
    
    var_cell = ws_cw.cell(row=row, column=6)
    var_cell.value = f'=E{row}-D{row}'
    var_cell.border = border
    
    pct_cell = ws_cw.cell(row=row, column=7)
    pct_cell.value = f'=IF(D{row}>0,E{row}/D{row}*100,0)'
    pct_cell.border = border
    pct_cell.number_format = '0.0'
    
    row += 1

cw_end = row - 1

ws_cw.column_dimensions['A'].width = 6
ws_cw.column_dimensions['B'].width = 32
ws_cw.column_dimensions['C'].width = 8
ws_cw.column_dimensions['D'].width = 12
ws_cw.column_dimensions['E'].width = 12
ws_cw.column_dimensions['F'].width = 12
ws_cw.column_dimensions['G'].width = 14

# COMBINED CHART - All 3 in one
cw_chart_start = row + 2
ws_cw.cell(row=cw_chart_start, column=1, value="Activity")
ws_cw.cell(row=cw_chart_start, column=2, value="Planned")
ws_cw.cell(row=cw_chart_start, column=3, value="Achieved")
ws_cw.cell(row=cw_chart_start, column=4, value="Variance")

for col in range(1, 5):
    ws_cw.cell(row=cw_chart_start, column=col).font = Font(bold=True, color='FFFFFF')
    ws_cw.cell(row=cw_chart_start, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_cw.cell(row=cw_chart_start, column=col).alignment = Alignment(horizontal='center')

for idx, r in compound_wall.iterrows():
    variance = r['Achieved'] - r['Planned']
    ws_cw.cell(row=cw_chart_start + idx + 1, column=1, value=r['Activity'][:22])
    ws_cw.cell(row=cw_chart_start + idx + 1, column=2, value=r['Planned'])
    ws_cw.cell(row=cw_chart_start + idx + 1, column=3, value=r['Achieved'])
    ws_cw.cell(row=cw_chart_start + idx + 1, column=4, value=variance)

chart_cw = BarChart()
chart_cw.type = "col"
chart_cw.style = 10
chart_cw.title = "Compound Wall - Planned vs Achieved vs Variance"
chart_cw.y_axis.title = "Quantity"
chart_cw.x_axis.title = ""
chart_cw.grouping = "clustered"
chart_cw.overlap = -30

data_cw = Reference(ws_cw, min_col=2, min_row=cw_chart_start, max_row=cw_chart_start + len(compound_wall), max_col=4)
cats_cw = Reference(ws_cw, min_col=1, min_row=cw_chart_start+1, max_row=cw_chart_start + len(compound_wall))
chart_cw.add_data(data_cw, titles_from_data=True)
chart_cw.set_categories(cats_cw)
chart_cw.width = 20
chart_cw.height = 12

chart_cw.series[0].graphicalProperties.solidFill = "5B9BD5"
chart_cw.series[1].graphicalProperties.solidFill = "70AD47"
chart_cw.series[2].graphicalProperties.solidFill = "C00000"

chart_cw.dataLabels = DataLabelList()
chart_cw.dataLabels.showVal = True

ws_cw.add_chart(chart_cw, f"A{cw_chart_start + len(compound_wall) + 2}")

# ============================================================
# KPI SUMMARY SHEET
# ============================================================

ws_kpi = wb.create_sheet("KPI Summary")

ws_kpi['A1'] = "KPI DASHBOARD - JKD 358 Production Block"
ws_kpi['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_kpi['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws_kpi.merge_cells('A1:F1')
ws_kpi['A1'].alignment = Alignment(horizontal='center')
ws_kpi.row_dimensions[1].height = 25

ws_kpi['A2'] = "Monthly Progress: February 2026"
ws_kpi['A2'].font = Font(size=11, italic=True)
ws_kpi.merge_cells('A2:F2')
ws_kpi['A2'].alignment = Alignment(horizontal='center')

# MAIN BUILDING KPI SECTION
ws_kpi['A4'] = "MAIN FACTORY BUILDING - INDIVIDUAL KPIs"
ws_kpi['A4'].font = Font(bold=True, size=12, color='FFFFFF')
ws_kpi['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_kpi.merge_cells('A4:E4')
ws_kpi['A4'].alignment = Alignment(horizontal='center')

kpi_headers = ['Activity', 'Unit', 'Planned', 'Achieved', 'Achievement %']
for col, h in enumerate(kpi_headers, 1):
    ws_kpi.cell(row=5, column=col, value=h).font = Font(bold=True, color='FFFFFF')
    ws_kpi.cell(row=5, column=col).fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    ws_kpi.cell(row=5, column=col).alignment = Alignment(horizontal='center')
    ws_kpi.cell(row=5, column=col).border = border

kpi_row = 6
for idx, r in main_building.iterrows():
    pct = round(r['Achieved']/r['Planned']*100, 1) if r['Planned']>0 else 0
    ws_kpi.cell(row=kpi_row, column=1, value=r['Activity'][:30]).border = border
    ws_kpi.cell(row=kpi_row, column=2, value=r['Unit']).border = border
    ws_kpi.cell(row=kpi_row, column=3, value=r['Planned']).border = border
    ws_kpi.cell(row=kpi_row, column=4, value=r['Achieved']).border = border
    
    pct_cell = ws_kpi.cell(row=kpi_row, column=5, value=pct).border = border
    ws_kpi.cell(row=kpi_row, column=5).number_format = '0.0'
    
    if pct >= 100:
        ws_kpi.cell(row=kpi_row, column=5).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif pct >= 80:
        ws_kpi.cell(row=kpi_row, column=5).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    else:
        ws_kpi.cell(row=kpi_row, column=5).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    kpi_row += 1

# COMPOUND WALL KPI SECTION
kpi_row += 1
ws_kpi.cell(row=kpi_row, column=1, value="COMPOUND WALL - INDIVIDUAL KPIs").font = Font(bold=True, size=12, color='FFFFFF')
ws_kpi.cell(row=kpi_row, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_kpi.merge_cells(f'A{kpi_row}:E{kpi_row}')
ws_kpi.cell(row=kpi_row, column=1).alignment = Alignment(horizontal='center')
kpi_row += 1

for col, h in enumerate(kpi_headers, 1):
    ws_kpi.cell(row=kpi_row, column=col, value=h).font = Font(bold=True, color='FFFFFF')
    ws_kpi.cell(row=kpi_row, column=col).fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    ws_kpi.cell(row=kpi_row, column=col).alignment = Alignment(horizontal='center')
    ws_kpi.cell(row=kpi_row, column=col).border = border
kpi_row += 1

for idx, r in compound_wall.iterrows():
    pct = round(r['Achieved']/r['Planned']*100, 1) if r['Planned']>0 else 0
    ws_kpi.cell(row=kpi_row, column=1, value=r['Activity'][:30]).border = border
    ws_kpi.cell(row=kpi_row, column=2, value=r['Unit']).border = border
    ws_kpi.cell(row=kpi_row, column=3, value=r['Planned']).border = border
    ws_kpi.cell(row=kpi_row, column=4, value=r['Achieved']).border = border
    
    pct_cell = ws_kpi.cell(row=kpi_row, column=5, value=pct).border = border
    ws_kpi.cell(row=kpi_row, column=5).number_format = '0.0'
    
    if pct >= 100:
        ws_kpi.cell(row=kpi_row, column=5).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif pct >= 80:
        ws_kpi.cell(row=kpi_row, column=5).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    else:
        ws_kpi.cell(row=kpi_row, column=5).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    kpi_row += 1

ws_kpi.column_dimensions['A'].width = 35
ws_kpi.column_dimensions['B'].width = 10
ws_kpi.column_dimensions['C'].width = 14
ws_kpi.column_dimensions['D'].width = 14
ws_kpi.column_dimensions['E'].width = 16

# ============================================================
# OUTPUT FILENAME
# ============================================================

output_path = r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\Monthly_Progress_Feb_2026.xlsx'

wb.save(output_path)
print(f"Excel saved: {output_path}")

mb_pct = round(main_building['Achieved'].sum()/main_building['Planned'].sum()*100, 1) if main_building['Planned'].sum()>0 else 0
cw_pct = round(compound_wall['Achieved'].sum()/compound_wall['Planned'].sum()*100, 1) if compound_wall['Planned'].sum()>0 else 0

print(f"\n=== SUMMARY ===")
print(f"Main Building - Planned: {main_building['Planned'].sum():.2f}, Achieved: {main_building['Achieved'].sum():.2f}, %: {mb_pct}%")
print(f"Compound Wall - Planned: {compound_wall['Planned'].sum():.2f}, Achieved: {compound_wall['Achieved'].sum():.2f}, %: {cw_pct}%")
