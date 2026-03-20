import os
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

excel_path = r"D:\JKD_Folder\JKD-PROJECT SITE\JKD_Weekly_Progrees\3.March_26\1March_Photos.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    print("Loaded workbook.", wb.sheetnames)
    ws = wb.active
    folder_path = r"D:\JKD_Folder\JKD-PROJECT SITE\JKD_SITE_PHOTOS\WPR_09-2-26\28-2-26"
    photo_name = os.listdir(folder_path)[0]
    photo_path = os.path.join(folder_path, photo_name)
    print("Photo path:", photo_path)
    
    img = OpenpyxlImage(photo_path)
    img.width = 454
    img.height = 680
    ws.add_image(img, "B2")
    ws.cell(row=2, column=1, value="Test from load_workbook")
    wb.save(excel_path)
    print("Saved workbook.")
except Exception as e:
    import traceback
    traceback.print_exc()
