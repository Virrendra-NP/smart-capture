import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, ttk
import win32com.client as win32

# --- STYLING ---
PRIMARY_COLOR = "1e293b" # Slate 800
ACCENT_COLOR = "3b82f6"  # Blue 500
SUCCESS_COLOR = "10b981" # Green
WARNING_COLOR = "f59e0b" # Yellow
DANGER_COLOR = "ef4444"  # Red
BG_COLOR = "#f8fafc"

def to_dt(val):
    if pd.isna(val) or val == '-': return None
    if isinstance(val, datetime): return val
    try: return pd.to_datetime(val)
    except: return None

def draw_kpi_card(ws, start_row, start_col, title, value, subtitle, bg_color="FFFFFF", font_color="334155"):
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
    t = ws.cell(row=start_row, column=start_col, value=title)
    t.font = Font(bold=True, color="64748B", size=10)
    t.alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=start_col + 3)
    v = ws.cell(row=start_row + 1, column=start_col, value=value)
    v.font = Font(bold=True, size=24, color=font_color)
    v.alignment = Alignment(horizontal='center', vertical='center')
    inner_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    for r in range(start_row + 1, start_row + 3):
        for c in range(start_col, start_col + 4):
            ws.cell(row=r, column=c).fill = inner_fill

    ws.merge_cells(start_row=start_row + 3, start_column=start_col, end_row=start_row + 3, end_column=start_col + 3)
    s = ws.cell(row=start_row + 3, column=start_col, value=subtitle)
    s.font = Font(size=9, color="94a3b8")
    s.alignment = Alignment(horizontal='center')
    
    side = Side(style='thin', color="e2e8f0")
    for r in range(start_row, start_row + 4):
        for c in range(start_col, start_col + 4):
            ws.cell(row=r, column=c).border = Border(left=side, right=side, top=side, bottom=side)

class ScheduleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("JKD - Schedule Engine")
        self.root.geometry("450x300")
        self.root.configure(bg=BG_COLOR)
        self.root.attributes("-topmost", True)
        
        self.source_path = r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Mar_DPR\Weekly_Dashboard_JKD - DPR - 10.03.2026.xlsx'
        self.out_path = r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Mar_DPR\Schedule_Analysis_Dashboard.xlsx'
        
        tk.Label(root, text="JKD SCHEDULE ANALYTICS ENGINE", font=("Segoe UI", 12, "bold"), bg=BG_COLOR, fg=f"#{PRIMARY_COLOR}").pack(pady=20)
        
        tk.Label(root, text="Click below after making changes in main sheet:", font=("Segoe UI", 9), bg=BG_COLOR).pack()
        
        self.update_btn = tk.Button(root, text="UPDATE DASHBOARD & REFRESH", bg=f"#{ACCENT_COLOR}", fg="white", 
                                    font=("Segoe UI", 11, "bold"), padx=20, pady=10, command=self.update_report)
        self.update_btn.pack(pady=20)
        
        tk.Label(root, text="Status: Ready", font=("Segoe UI", 8), bg=BG_COLOR, fg="#64748B").pack(side="bottom", pady=10)

    def update_report(self):
        try:
            self.update_btn.config(state="disabled", text="UPDATING...")
            self.root.update()
            
            # --- CLOSE FILE IF OPEN (via win32) ---
            try:
                xl = win32.GetActiveObject("Excel.Application")
                for wb in xl.Workbooks:
                    if wb.FullName == self.out_path:
                        wb.Save()
                        wb.Close()
            except: pass # Excel not open or file not open
            
            # --- ANALYSIS LOGIC ---
            report_date = datetime(2026, 3, 10)
            df = pd.read_excel(self.source_path, sheet_name='Sheet1', header=0)
            
            results = []
            counts = {"Completed": 0, "On Track": 0, "Delayed": 0}
            
            for _, row in df.iterrows():
                try:
                    id_val = row.get('ID')
                    if pd.isna(id_val) or not str(id_val).strip().isdigit(): continue
                    
                    task = row.get('Task Name', 'Unknown')
                    b_start = to_dt(row.get('Baseline Start'))
                    b_finish = to_dt(row.get('Baseline Finish'))
                    a_start = to_dt(row.get('Actual Start'))
                    a_finish = to_dt(row.get('Actual Finish'))
                    f_finish = to_dt(row.get('Finish'))
                    pct = row.get('% Complete', 0)
                    
                    status = "On Track"
                    s_delay = 0
                    f_delay = 0
                    
                    # Logic for START
                    if a_start and b_start:
                        s_delay = (a_start - b_start).days
                    
                    # Logic for FINISH
                    target_finish = a_finish if pct >= 1 else f_finish
                    if target_finish and b_finish:
                        f_delay = (target_finish - b_finish).days
                    
                    # Overall Status
                    if pct >= 1:
                        status = "Completed"
                        counts["Completed"] += 1
                    elif f_delay > 0 or s_delay > 0:
                        status = "Delayed"
                        counts["Delayed"] += 1
                    else:
                        status = "On Track"
                        counts["On Track"] += 1
                    
                    results.append({
                        'ID': id_val, 
                        'Task': task, 
                        'B_Start': b_start,
                        'A_Start': a_start,
                        'S_Delay': s_delay,
                        'B_Finish': b_finish, 
                        'A_Finish': target_finish,
                        'F_Delay': f_delay,
                        'Progress': f"{round(pct * 100)}%", 
                        'Status': status
                    })
                except Exception as e:
                    print(f"Row error: {e}")
                    continue

            # --- EXCEL GENERATION ---
            wb = Workbook()
            ws = wb.active
            ws.title = "SCHEDULE DASHBOARD"
            ws.sheet_view.showGridLines = False
            
            # Dash UI
            ws.merge_cells("B2:P5")
            h_cell = ws["B2"]; h_cell.value = "JKD PROJECT SCHEDULE ANALYTICS"
            h_cell.font = Font(bold=True, size=24, color="FFFFFF")
            h_cell.fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
            h_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            draw_kpi_card(ws, 9, 2, "COMPLETED", counts["Completed"], "Total Project Tasks", "f1f5f9", SUCCESS_COLOR)
            draw_kpi_card(ws, 9, 7, "ON TRACK", counts["On Track"], "Ongoing Progress", "f1f5f9", ACCENT_COLOR)
            draw_kpi_card(ws, 9, 12, "DELAYED", counts["Delayed"], "Issues Detected", "fef2f2", DANGER_COLOR)

            # Table for Detailed Comparison
            ws.merge_cells("B15:P15")
            ws["B15"].value = "COMPLETE SCHEDULE TRACKING: BASELINE VS ACTUAL (START & FINISH)"; ws["B15"].font = Font(bold=True, size=14, color=PRIMARY_COLOR)
            
            heads = ["ID", "Activity Name", "B.Start", "A.Start", "S.Var", "B.Finish", "A/F.Finish", "F.Var", "Progress", "Status"]
            for i, h in enumerate(heads):
                c = ws.cell(row=17, column=2 + i, value=h)
                c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
                c.alignment = Alignment(horizontal='center')
                c.border = Border(left=Side(style='thin', color="FFFFFF"), right=Side(style='thin', color="FFFFFF"))

            curr_row = 18
            stripe_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
            red_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            green_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
            
            for i, r in enumerate(results[:50]): # Top 50
                row_fill = stripe_fill if i % 2 == 1 else None
                row_data = [
                    r['ID'], 
                    r['Task'][:50],
                    r['B_Start'].strftime('%d-%b') if r['B_Start'] else '-',
                    r['A_Start'].strftime('%d-%b') if r['A_Start'] else '-',
                    r['S_Delay'],
                    r['B_Finish'].strftime('%d-%b') if r['B_Finish'] else '-',
                    r['A_Finish'].strftime('%d-%b') if r['A_Finish'] else '-',
                    r['F_Delay'],
                    r['Progress'],
                    r['Status']
                ]
                
                for col_idx, val in enumerate(row_data):
                    c = ws.cell(row=curr_row, column=2 + col_idx, value=val)
                    if row_fill: c.fill = row_fill
                    c.border = Border(bottom=Side(style='thin', color="cbd5e1"))
                    if col_idx in [4, 7] and isinstance(val, int): # Variation columns
                        if val > 0: c.font = Font(color=DANGER_COLOR, bold=True)
                        elif val < 0: c.font = Font(color=SUCCESS_COLOR, bold=True)
                curr_row += 1

            # Conditional Formatting for Status column
            stat_col = "K"
            ws.conditional_formatting.add(f'{stat_col}18:{stat_col}{curr_row-1}', CellIsRule(operator='equal', formula=['"Delayed"'], fill=red_fill))
            ws.conditional_formatting.add(f'{stat_col}18:{stat_col}{curr_row-1}', CellIsRule(operator='equal', formula=['"Completed"'], fill=green_fill))
            
            # Column Sizing
            ws.column_dimensions['C'].width = 45
            for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws.column_dimensions[col].width = 12

            # --- SEPARATE SHEET FOR INCOMPLETE TASKS (LIVE DELAY ANALYSIS) ---
            ws_live = wb.create_sheet("LIVE DELAY ANALYSIS")
            ws_live.sheet_view.showGridLines = False
            
            # Header
            ws_live.merge_cells("A1:G2")
            l_cell = ws_live["A1"]; l_cell.value = "INCOMPLETE TASKS: LIVE PROGRESS vs BASELINE OVERVIEW"
            l_cell.font = Font(bold=True, size=18, color="FFFFFF")
            l_cell.fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
            l_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            today_now = datetime.now()
            ws_live.merge_cells("A3:G3")
            ws_live["A3"] = f"ANALYSIS AS OF TODAY: {today_now.strftime('%d %b %Y')}"
            ws_live["A3"].font = Font(italic=True, color="64748B")
            ws_live["A3"].alignment = Alignment(horizontal='center')

            l_heads = ["ID", "Task Name", "Baseline Finish", "Forecast Finish", "Status vs Today", "Days Overdue", "Expected Variance"]
            for i, h in enumerate(l_heads):
                c = ws_live.cell(row=5, column=i+1, value=h)
                c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
                c.alignment = Alignment(horizontal='center')

            l_row = 6
            t_date = today_now.date()
            for i, r in enumerate([res for res in results if res['Status'] != "Completed"]):
                # Zebra striping
                r_fill = stripe_fill if i % 2 == 1 else None
                
                b_fin = r['B_Finish'].date() if r['B_Finish'] else None
                v_today = "ON TRACK"
                ov_days = 0
                if b_fin and t_date > b_fin:
                    v_today = "OVERDUE (CRITICAL)"
                    ov_days = (t_date - b_fin).days
                
                curr_vals = [
                    r['ID'], 
                    r['Task'], 
                    r['B_Finish'].strftime('%d-%b-%Y') if r['B_Finish'] else '-', 
                    r['A_Finish'].strftime('%d-%b-%Y') if r['A_Finish'] else '-',
                    v_today,
                    f"{ov_days} Days" if ov_days > 0 else "-",
                    f"{r['F_Delay']} Days"
                ]
                
                for col_idx, val in enumerate(curr_vals):
                    cel = ws_live.cell(row=l_row, column=col_idx+1, value=val)
                    if r_fill: cel.fill = r_fill
                    cel.border = Border(bottom=Side(style='thin', color="cbd5e1"))
                    
                    if col_idx == 4 and v_today.startswith("OVERDUE"): cel.font = Font(color=DANGER_COLOR, bold=True)
                    if col_idx == 6 and r['F_Delay'] > 0: cel.font = Font(color=DANGER_COLOR, bold=True)
                
                l_row += 1
            
            ws_live.column_dimensions['B'].width = 50
            for col_l in ['C', 'D', 'E', 'F', 'G']: ws_live.column_dimensions[col_l].width = 18

            wb.save(self.out_path)
            os.startfile(self.out_path)
            messagebox.showinfo("Success", "Schedule Dashboard Updated Successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed: {str(e)}")
        finally:
            self.update_btn.config(state="normal", text="UPDATE DASHBOARD & REFRESH")

if __name__ == "__main__":
    root = tk.Tk()
    app = ScheduleApp(root)
    root.mainloop()

