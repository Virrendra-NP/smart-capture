import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

df = pd.read_excel(r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\JKD - DPR - 28.02.2026 (1) (1) (1).xlsx', sheet_name="Feb'26", header=None)

date_cols = {41: '21-Feb', 43: '23-Feb', 45: '24-Feb', 47: '25-Feb', 49: '26-Feb', 51: '27-Feb', 53: '28-Feb'}
date_full = {41: '21-Feb-2026', 43: '23-Feb-2026', 45: '24-Feb-2026', 47: '25-Feb-2026', 49: '26-Feb-2026', 51: '27-Feb-2026', 53: '28-Feb-2026'}

def get_weekly_data(start_row, end_row):
    data_rows = []
    for i in range(start_row, end_row):
        activity = df.iloc[i, 1]
        unit = df.iloc[i, 2]
        if pd.notna(activity) and activity not in ['Main Factory Building', 'Compound Wall'] and not isinstance(activity, float):
            p, a = 0, 0
            for col in date_cols.keys():
                p += float(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else 0
                a += float(df.iloc[i, col + 1]) if pd.notna(df.iloc[i, col + 1]) else 0
            if p > 0 or a > 0:
                data_rows.append({'Activity': activity, 'Unit': unit, 'Planned': p, 'Achieved': a, '%': round((a/p*100) if p>0 else 0, 1)})
    return pd.DataFrame(data_rows)

main_building = get_weekly_data(7, 26)
main_building['Category'] = 'Main Factory Building'

compound_wall = get_weekly_data(27, 40)
compound_wall['Category'] = 'Compound Wall'

all_data = pd.concat([main_building, compound_wall], ignore_index=True)

wb = Workbook()
ws = wb.active
ws.title = "Weekly Progress Report"

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(bold=True, size=14)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.merge_cells('A1:F1')
ws['A1'] = "JKD 358 - Weekly Progress Report (21-Feb-2026 to 28-Feb-2026)"
ws['A1'].font = Font(bold=True, size=14, color='2F5496')
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:F2')
ws['A2'] = "Client: JK Defence and Aerospace | Project: JKD 358 (Production Block)"
ws['A2'].font = Font(size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

ws['A4'] = "MAIN FACTORY BUILDING"
ws['A4'].font = Font(bold=True, size=12, color='FFFFFF')
ws['A4'].fill = header_fill
ws.merge_cells('A4:F4')

headers = ['Activity', 'Unit', 'Planned', 'Achieved', 'Achievement %', 'Status']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col, value=h)
    cell.font = header_font
    cell.fill = subheader_fill
    cell.border = border

row = 6
for _, r in main_building.iterrows():
    ws.cell(row=row, column=1, value=r['Activity']).border = border
    ws.cell(row=row, column=2, value=r['Unit']).border = border
    ws.cell(row=row, column=3, value=r['Planned']).border = border
    ws.cell(row=row, column=4, value=r['Achieved']).border = border
    pct = ws.cell(row=row, column=5, value=r['%'])
    pct.border = border
    
    status = ws.cell(row=row, column=6)
    status.border = border
    if r['%'] >= 100:
        pct.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        status.value = 'Excellent'
        status.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif r['%'] >= 80:
        pct.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        status.value = 'On Track'
        status.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    else:
        pct.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        status.value = 'Needs Attention'
        status.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    row += 1

ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=row, column=3, value=main_building['Planned'].sum()).font = Font(bold=True)
ws.cell(row=row, column=4, value=main_building['Achieved'].sum()).font = Font(bold=True)
total_pct = ws.cell(row=row, column=5, value=round(main_building['Achieved'].sum()/main_building['Planned'].sum()*100, 1) if main_building['Planned'].sum()>0 else 0)
total_pct.font = Font(bold=True)
row += 2

ws.cell(row=row, column=1, value="COMPOUND WALL")
ws.cell(row=row, column=1).font = Font(bold=True, size=12, color='FFFFFF')
ws.cell(row=row, column=1).fill = header_fill
ws.merge_cells(f'A{row}:F{row}')
row += 1

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = subheader_fill
    cell.border = border
row += 1

for _, r in compound_wall.iterrows():
    ws.cell(row=row, column=1, value=r['Activity']).border = border
    ws.cell(row=row, column=2, value=r['Unit']).border = border
    ws.cell(row=row, column=3, value=r['Planned']).border = border
    ws.cell(row=row, column=4, value=r['Achieved']).border = border
    pct = ws.cell(row=row, column=5, value=r['%'])
    pct.border = border
    
    status = ws.cell(row=row, column=6)
    status.border = border
    if r['%'] >= 100:
        pct.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        status.value = 'Excellent'
        status.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif r['%'] >= 80:
        pct.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        status.value = 'On Track'
        status.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    else:
        pct.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        status.value = 'Needs Attention'
        status.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    row += 1

ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=row, column=3, value=compound_wall['Planned'].sum()).font = Font(bold=True)
ws.cell(row=row, column=4, value=compound_wall['Achieved'].sum()).font = Font(bold=True)
total_pct2 = ws.cell(row=row, column=5, value=round(compound_wall['Achieved'].sum()/compound_wall['Planned'].sum()*100, 1) if compound_wall['Planned'].sum()>0 else 0)
total_pct2.font = Font(bold=True)
row += 2

ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 18

ws2 = wb.create_sheet("KPI Dashboard")

ws2['A1'] = "KPI DASHBOARD - Weekly Progress (21-Feb-2026 to 28-Feb-2026)"
ws2['A1'].font = Font(bold=True, size=16, color='2F5496')
ws2.merge_cells('A1:E1')

mb_planned = main_building['Planned'].sum()
mb_achieved = main_building['Achieved'].sum()
cw_planned = compound_wall['Planned'].sum()
cw_achieved = compound_wall['Achieved'].sum()
total_planned = mb_planned + cw_planned
total_achieved = mb_achieved + cw_achieved
overall_pct = round(total_achieved/total_planned*100, 1) if total_planned>0 else 0
mb_pct = round(mb_achieved/mb_planned*100, 1) if mb_planned>0 else 0
cw_pct = round(cw_achieved/cw_planned*100, 1) if cw_planned>0 else 0

kpis = [
    ('Total Planned', round(total_planned, 2)),
    ('Total Achieved', round(total_achieved, 2)),
    ('Overall Achievement', f"{overall_pct}%"),
    ('Main Building Achievement', f"{mb_pct}%"),
    ('Compound Wall Achievement', f"{cw_pct}%"),
]

kpi_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
for i, (label, value) in enumerate(kpis, 3):
    ws2.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
    ws2.cell(row=i, column=2, value=value).font = Font(size=14, bold=True, color='2F5496')
    ws2.cell(row=i, column=1).fill = kpi_fill
    ws2.cell(row=i, column=2).fill = kpi_fill
    ws2.cell(row=i, column=1).border = border
    ws2.cell(row=i, column=2).border = border

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 18

output_path = r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\Weekly_Progress_Report_21-28_Feb_2026.xlsx'
wb.save(output_path)
print(f"Excel saved: {output_path}")
print(f"\n=== SUMMARY ===")
print(f"Main Building - Planned: {mb_planned:.2f}, Achieved: {mb_achieved:.2f}")
print(f"Compound Wall - Planned: {cw_planned:.2f}, Achieved: {cw_achieved:.2f}")
print(f"Overall - Planned: {total_planned:.2f}, Achieved: {total_achieved:.2f}")
