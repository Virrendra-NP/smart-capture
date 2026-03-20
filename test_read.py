import openpyxl

excel_path = r"D:\JKD_Folder\JKD-PROJECT SITE\JKD_Weekly_Progrees\3.March_26\1March_Photos.xlsx"
try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    print("A1:", ws["A1"].value)
    print("B1:", ws["B1"].value)
    print("A2:", ws["A2"].value)
    print("B2:", ws["B2"].value)
    print("A3:", ws["A3"].value)
    
except Exception as e:
    print("Error:", e)
