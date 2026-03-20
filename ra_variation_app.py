import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
import traceback

def create_variation_statement():
    bill_path = file_var.get()
    source_sheet_name = sheet_var.get()
    
    if not bill_path or not os.path.exists(bill_path):
        messagebox.showerror("Error", "Please select a valid RA Bill file.")
        return
        
    if not source_sheet_name:
        messagebox.showerror("Error", "Please select the Source Sheet name.")
        return

    btn.config(state="disabled", text="Processing...")
    root.update()

    try:
        wb = openpyxl.load_workbook(bill_path)
        
        if source_sheet_name not in wb.sheetnames:
            messagebox.showerror("Error", f"Sheet '{source_sheet_name}' not found.")
            btn.config(state="normal", text="ADD VARIATION STATEMENT SHEET")
            return
            
        target_name = "Variation Statement"
        if target_name in wb.sheetnames:
            wb.remove(wb[target_name])
            
        ws_var = wb.create_sheet(target_name)
        ws_src = wb[source_sheet_name]
        
        # --- STYLES ---
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        sub_header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrapText=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Color Fills for Status
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light Red (Addition)
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow (Near BOQ)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green (Below BOQ)

        # Number format: Hide ZEROS
        # Format: Positive; Negative; Zero; Text
        HIDE_ZERO_FORMAT = '#,##0.00;-#,##0.00;;@'

        # --- HEADERS ---
        ws_var.merge_cells('A12:A13'); ws_var['A12'] = "Sl.No"
        ws_var.merge_cells('B12:B13'); ws_var['B12'] = "Description of Work"
        ws_var.merge_cells('C12:F12'); ws_var['C12'] = "As per BOQ"
        ws_var.merge_cells('G12:H12'); ws_var['G12'] = "Cumulative Upto \nPrevious Bill"
        ws_var.merge_cells('I12:J12'); ws_var['I12'] = "Cumulative Upto \nPresent Bill"
        ws_var.merge_cells('K12:L12'); ws_var['K12'] = "ADDITION"
        ws_var.merge_cells('M12:N12'); ws_var['M12'] = "OMMISSION"
        
        sub_headers = {
            'C': "Unit", 'D': "Qty", 'E': "Rate", 'F': "Amount",
            'G': "Qty", 'H': "Amount",
            'I': "Qty", 'J': "Amount",
            'K': "Qty", 'L': "Amount",
            'M': "Qty", 'N': "Amount"
        }
        
        for col, title in sub_headers.items():
            ws_var[f"{col}13"] = title
            
        for r_idx in [12, 13]:
            for c_idx in range(1, 15): 
                cell = ws_var.cell(row=r_idx, column=c_idx)
                cell.font = white_font
                cell.fill = header_fill if r_idx == 12 else sub_header_fill
                cell.alignment = center_align
                cell.border = border

        # --- DATA PROCESSING ---
        var_row = 14
        max_src_row = ws_src.max_row
        s_name = source_sheet_name
        
        for src_row in range(13, max_src_row + 1):
            sl_no_raw = ws_src.cell(row=src_row, column=1).value
            desc_raw = ws_src.cell(row=src_row, column=2).value
            
            if desc_raw or sl_no_raw:
                ws_var.cell(row=var_row, column=1, value=f"='{s_name}'!A{src_row}")
                ws_var.cell(row=var_row, column=2, value=f"='{s_name}'!B{src_row}")
                ws_var.cell(row=var_row, column=3, value=f"='{s_name}'!C{src_row}")
                ws_var.cell(row=var_row, column=4, value=f"='{s_name}'!I{src_row}")
                ws_var.cell(row=var_row, column=5, value=f"='{s_name}'!J{src_row}")
                ws_var.cell(row=var_row, column=6, value=f"=D{var_row}*E{var_row}")
                
                ws_var.cell(row=var_row, column=7, value=f"='{s_name}'!L{src_row}")
                ws_var.cell(row=var_row, column=8, value=f"='{s_name}'!M{src_row}")
                
                ws_var.cell(row=var_row, column=9, value=f"='{s_name}'!L{src_row}+'{s_name}'!N{src_row}")
                ws_var.cell(row=var_row, column=10, value=f"=I{var_row}*E{var_row}")
                
                ws_var.cell(row=var_row, column=11, value=f"=IF(I{var_row}>D{var_row},I{var_row}-D{var_row},0)")
                ws_var.cell(row=var_row, column=12, value=f"=K{var_row}*E{var_row}")
                
                ws_var.cell(row=var_row, column=13, value=f"=IF(I{var_row}<D{var_row},D{var_row}-I{var_row},0)")
                ws_var.cell(row=var_row, column=14, value=f"=M{var_row}*E{var_row}")
                
                # --- APPLY FORMATS ---
                for c_idx in range(1, 15):
                    cell = ws_var.cell(row=var_row, column=c_idx)
                    cell.border = border
                    if c_idx >= 4:
                        cell.number_format = HIDE_ZERO_FORMAT
                
                var_row += 1

        # --- CONDITIONAL FORMATTING RULES ---
        # Addition (Red) if Cumulative Qty (I) > BOQ Qty (D)
        # Apply to description row for high visibility
        data_range = f"A14:N{var_row-1}"
        
        # Rule 1: Addition (Red) -> if I > D
        ws_var.conditional_formatting.add(data_range, FormulaRule(formula=[f'=$I14>$D14'], fill=red_fill))
        
        # Rule 2: Near BOQ (Yellow) -> if I >= 0.9 * D AND I <= D
        ws_var.conditional_formatting.add(data_range, FormulaRule(formula=[f'AND($I14<=$D14, $I14>=$D14*0.9, $D14>0)'], fill=yellow_fill))
        
        # Rule 3: Omission/Below (Green) -> if I < 0.9 * D
        ws_var.conditional_formatting.add(data_range, FormulaRule(formula=[f'AND($I14<$D14*0.9, $D14>0)'], fill=green_fill))

        # Adjust widths
        ws_var.column_dimensions['A'].width = 8
        ws_var.column_dimensions['B'].width = 60
        ws_var.column_dimensions['C'].width = 10
        for col in "DEFGHIJKLMN":
            ws_var.column_dimensions[col].width = 16

        wb.save(bill_path)
        messagebox.showinfo("Success", f"Variation Statement has been updated!\n- Red: Addition\n- Yellow: Near BOQ (>90%)\n- Green: Omission/Below\n- All '0.00' values are hidden.")
        os.startfile(bill_path)

    except PermissionError:
        messagebox.showerror("Error", f"Could not save file! Please close the Excel file first.")
    except Exception as e:
        err_msg = traceback.format_exc()
        with open("ra_error_log.txt", "w") as f:
            f.write(err_msg)
        messagebox.showerror("Error", f"Error: {str(e)}\nSee 'ra_error_log.txt'.")
    finally:
        btn.config(state="normal", text="ADD VARIATION STATEMENT SHEET")

# GUI UI
def select_file():
    path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if path:
        file_var.set(path)
        try:
            wb_t = openpyxl.load_workbook(path, read_only=True)
            sheet_combo['values'] = wb_t.sheetnames
            if wb_t.sheetnames: sheet_var.set(wb_t.sheetnames[0])
            wb_t.close()
        except: pass

root = tk.Tk()
root.title("JKD RA Bill - Variation Statement Tool v2")
root.geometry("650x350")
root.eval('tk::PlaceWindow . center')

tk.Label(root, text="Step 1: Select your RA Bill Excel File:", font=("Arial", 10, "bold")).pack(pady=(20, 5))
file_var = tk.StringVar()
frame_file = tk.Frame(root)
frame_file.pack(fill="x", padx=30)
tk.Entry(frame_file, textvariable=file_var, width=65).pack(side="left", padx=(0, 10))
tk.Button(frame_file, text="Browse", command=select_file).pack(side="left")

tk.Label(root, text="Step 2: Select Sheet with RA Data (Header in Row 12):", font=("Arial", 10, "bold")).pack(pady=(20, 5))
sheet_var = tk.StringVar()
sheet_combo = ttk.Combobox(root, textvariable=sheet_var, width=45)
sheet_combo.pack()

btn = tk.Button(root, text="ADD VARIATION STATEMENT SHEET", command=create_variation_statement, 
               bg="#2F75B5", fg="white", font=("Arial", 11, "bold"), padx=20, pady=12)
btn.pack(pady=30)

root.mainloop()
