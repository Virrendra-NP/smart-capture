import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
from tkcalendar import DateEntry

# --- UI STYLING CONSTANTS ---
PRIMARY_COLOR = "#1e293b"  # Slate 800
ACCENT_COLOR = "#3b82f6"   # Blue 500
BG_COLOR = "#f8fafc"       # Slate 50
SUCCESS_COLOR = "#10b981"  # Emerald 500
WARNING_COLOR = "#f59e0b"  # Amber 500
DANGER_COLOR = "#ef4444"   # Red 500

def draw_kpi_card(ws, start_row, start_col, title, value, subtitle, bg_color="FFFFFF", font_color="334155", num_format="0.0"):
    """Draws a premium looking KPI card in the Excel sheet."""
    # Outer container border (using merged cells)
    card_width = 4
    card_height = 4
    
    # Title row
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + card_width - 1)
    t = ws.cell(row=start_row, column=start_col, value=title)
    t.font = Font(bold=True, color="64748B", size=10)
    t.alignment = Alignment(horizontal='center', vertical='bottom')
    
    # Value row (Main Metric)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=start_col + card_width - 1)
    v = ws.cell(row=start_row + 1, column=start_col, value=value)
    v.font = Font(bold=True, size=24, color=font_color)
    v.alignment = Alignment(horizontal='center', vertical='center')
    v.number_format = num_format
    inner_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    
    for r in range(start_row + 1, start_row + 3):
        for c in range(start_col, start_col + card_width):
            ws.cell(row=r, column=c).fill = inner_fill

    # Subtitle row
    ws.merge_cells(start_row=start_row + 3, start_column=start_col, end_row=start_row + 3, end_column=start_col + card_width - 1)
    s = ws.cell(row=start_row + 3, column=start_col, value=subtitle)
    s.font = Font(size=9, color="94a3b8")
    s.alignment = Alignment(horizontal='center', vertical='top')
    
    # Border
    thin_border = Side(style='thin', color="e2e8f0")
    for r in range(start_row, start_row + 4):
        for c_idx in range(start_col, start_col + card_width):
            cell = ws.cell(row=r, column=c_idx)
            left = thin_border if c_idx == start_col else None
            right = thin_border if c_idx == start_col + card_width - 1 else None
            top = thin_border if r == start_row else None
            bottom = thin_border if r == start_row + 3 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def create_premium_chart(ws_data, title, data_ref_params, cats_ref_params):
    """Creates a high-end horizontal bar chart for the dashboard."""
    chart = BarChart()
    chart.type = "bar" # Use Bar (horizontal) for better readability of names and top/bottom comparison
    chart.style = 10
    chart.title = title
    chart.x_axis.title = "Quantity"
    chart.y_axis.title = ""
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.gapWidth = 120
    
    # Data Reference
    data = Reference(ws_data, **data_ref_params)
    cats = Reference(ws_data, **cats_ref_params)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    chart.width = 38
    chart.height = 20
    
    # Modern Styling for Series
    colors = ["3b82f6", "10b981", "ef4444"] # Blue, Green, Red
    for i, s in enumerate(chart.series):
        s.graphicalProperties.solidFill = colors[i % len(colors)]
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True # Show numbers
        s.dLbls.showSerName = False # REMOVE "Planned/Achieved" text from bars
        s.dLbls.showCatName = False
        s.dLbls.position = "outEnd"
        
    # Axis Cleaning & Professional Look
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.majorGridlines = None 
    chart.legend.position = "t"
    
    # ADD DATA TABLE - Provides the exact "Planned Top / Achieved Bottom" grid requested
    chart.plot_area.dTable = DataTable()
    chart.plot_area.dTable.showHorzBorder = True
    chart.plot_area.dTable.showVertBorder = True
    chart.plot_area.dTable.showOutline = True
    chart.plot_area.dTable.showKeys = True # Show Legend Keys in the table
    
    return chart

class DashingDashboardApp:
    def __init__(self, root):
        self.root = root
        self.root.title("JKD - Unimaginable Progress Dashboard")
        self.root.geometry("850x700")
        self.root.configure(bg=BG_COLOR)
        
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("TFrame", background=BG_COLOR)
        self.style.configure("TLabel", background=BG_COLOR, font=("Segoe UI", 10))
        self.style.configure("Header.TLabel", font=("Segoe UI", 14, "bold"), foreground=PRIMARY_COLOR)
        
        self.input_file = tk.StringVar()
        self.output_file = tk.StringVar()
        self.sheet_name = tk.StringVar()
        
        self.date_cols = {}
        self.df_cache = None
        
        self.create_widgets()
        
    def create_widgets(self):
        # Header Area
        header_frame = tk.Frame(self.root, bg=PRIMARY_COLOR, height=80)
        header_frame.pack(fill="x")
        
        tk.Label(header_frame, text="JKD PROGRESS ANALYTICS ENGINE", bg=PRIMARY_COLOR, fg="white", 
                 font=("Segoe UI", 18, "bold")).pack(pady=20)
        
        main_container = ttk.Frame(self.root, padding="30")
        main_container.pack(fill="both", expand=True)
        
        # 1. File Selection
        sec1 = ttk.LabelFrame(main_container, text=" 1. Data Source ", padding=15)
        sec1.pack(fill="x", pady=(0, 20))
        
        ttk.Label(sec1, text="Excel DPR Path:").pack(anchor="w")
        f_row = ttk.Frame(sec1)
        f_row.pack(fill="x")
        ttk.Entry(f_row, textvariable=self.input_file).pack(side="left", fill="x", expand=True, padx=(0,10))
        ttk.Button(f_row, text="Browse", command=self.browse_input).pack(side="right")
        
        # 2. Sheet & Dates
        sec2 = ttk.LabelFrame(main_container, text=" 2. Period Analysis ", padding=15)
        sec2.pack(fill="x", pady=0)
        
        g_row = ttk.Frame(sec2)
        g_row.pack(fill="x", pady=(0, 15))
        ttk.Label(g_row, text="Sheet:").pack(side="left", padx=(0, 5))
        self.sheet_combo = ttk.Combobox(g_row, textvariable=self.sheet_name, width=20)
        self.sheet_combo.pack(side="left", padx=(0, 15))
        ttk.Button(g_row, text="Scan Sheet for Dates", command=self.load_dates_from_sheet).pack(side="left")
        
        d_row = ttk.Frame(sec2)
        d_row.pack(fill="x")
        
        # Start Date
        s_box = ttk.Frame(d_row)
        s_box.pack(side="left", fill="x", expand=True)
        ttk.Label(s_box, text="Analysis Start:").pack(anchor="w")
        self.start_date_entry = DateEntry(s_box, width=15, date_pattern='dd-mm-yyyy', showweeknumbers=False)
        self.start_date_entry.pack(anchor="w")
        
        # End Date
        e_box = ttk.Frame(d_row)
        e_box.pack(side="left", fill="x", expand=True)
        ttk.Label(e_box, text="Analysis End:").pack(anchor="w")
        self.end_date_entry = DateEntry(e_box, width=15, date_pattern='dd-mm-yyyy', showweeknumbers=False)
        self.end_date_entry.pack(anchor="w")
        
        # 3. Output
        sec3 = ttk.LabelFrame(main_container, text=" 3. Save Report ", padding=15)
        sec3.pack(fill="x", pady=20)
        
        o_row = ttk.Frame(sec3)
        o_row.pack(fill="x")
        ttk.Entry(o_row, textvariable=self.output_file).pack(side="left", fill="x", expand=True, padx=(0,10))
        ttk.Button(o_row, text="Select Path", command=self.browse_output).pack(side="right")
        
        # Action Button
        self.gen_btn = tk.Button(main_container, text="GENERATE UNIMAGINABLE DASHBOARD", bg=ACCENT_COLOR, fg="white",
                                font=("Segoe UI", 12, "bold"), command=self.generate, relief="flat", padx=20, pady=10)
        self.gen_btn.pack(pady=10)

    def browse_input(self):
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file:
            self.input_file.set(file)
            try:
                xl = pd.ExcelFile(file)
                self.sheet_combo['values'] = xl.sheet_names
                if "Feb'26" in xl.sheet_names: self.sheet_name.set("Feb'26")
                elif xl.sheet_names: self.sheet_name.set(xl.sheet_names[0])
            except: pass

    def load_dates_from_sheet(self):
        file = self.input_file.get()
        sheet = self.sheet_name.get()
        if not file or not sheet: return
        
        try:
            df = pd.read_excel(file, sheet_name=sheet, header=None)
            self.df_cache = df
            self.date_cols.clear()
            
            # Dates start from row 4 (index 4) columns 3 (D) to 52 (BA)
            # SAFETY CHECK: Ensure DataFrame has enough rows and columns
            num_rows, num_cols = df.shape
            
            if num_rows < 5:
                messagebox.showwarning("Warning", f"Sheet '{sheet}' only has {num_rows} rows. Date headers expected in row 5.")
                return

            for c in range(3, 53): 
                if c >= num_cols: break # Stop if we reach end of columns
                
                try:
                    val = df.iloc[4, c]
                    if pd.notna(val):
                        date_obj = None
                        if isinstance(val, datetime): date_obj = val.date()
                        else:
                            try: 
                                # Try parsing different formats
                                d_val = pd.to_datetime(val)
                                if pd.notna(d_val): date_obj = d_val.date()
                            except: continue
                        
                        if date_obj:
                            self.date_cols[c] = date_obj
                except IndexError:
                    break # Safety break
            
            if self.date_cols:
                all_dates = sorted(self.date_cols.values())
                self.start_date_entry.set_date(all_dates[0])
                self.end_date_entry.set_date(all_dates[-1])
                messagebox.showinfo("Success", f"Found {len(self.date_cols)} dates in range D5:BA5")
            else:
                messagebox.showwarning("Warning", "No valid dates found in row 5 (Columns D to BA)!")
        except Exception as e:
            messagebox.showerror("Sheet Analysis Error", f"Failed to scan sheet: {str(e)}\n\nPlease ensure row 5 contains dates.")

    def browse_output(self):
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file: self.output_file.set(file)

    def generate(self):
        start_d = self.start_date_entry.get_date()
        end_d = self.end_date_entry.get_date()
        out_path = self.output_file.get()
        
        if not out_path: 
            messagebox.showerror("Error", "Select output path first")
            return
            
        filtered_cols = {c: d for c, d in self.date_cols.items() if start_d <= d <= end_d}
        if not filtered_cols:
            messagebox.showerror("Error", "No dates in data match this period")
            return
            
        self.gen_btn.config(state="disabled", text="MAKING IT DASHING...")
        self.root.update()
        
        try:
            df = pd.read_excel(self.input_file.get(), sheet_name=self.sheet_name.get(), header=None)
            
            def aggregate_rows(start_row, end_row):
                results = []
                num_rows, num_cols = df.shape
                for i in range(start_row, end_row):
                    if i >= num_rows: break
                    
                    try:
                        act = df.iloc[i, 1]
                        unit = df.iloc[i, 2]
                        if pd.notna(act) and not isinstance(act, float):
                            p_sum, a_sum = 0, 0
                            for col_idx in filtered_cols.keys():
                                # Safety column checks
                                if col_idx < num_cols:
                                    pval = df.iloc[i, col_idx]
                                    p_sum += float(pval) if pd.notna(pval) else 0
                                if col_idx + 1 < num_cols:
                                    aval = df.iloc[i, col_idx + 1]
                                    a_sum += float(aval) if pd.notna(aval) else 0
                                    
                            if p_sum > 0 or a_sum > 0:
                                results.append({'Activity': str(act), 'Unit': str(unit), 'Planned': p_sum, 'Achieved': a_sum, 'Variance': a_sum - p_sum})
                    except Exception as loop_e:
                        print(f"Skipping row {i} due to Error: {loop_e}")
                        continue
                return pd.DataFrame(results)

            # Results containers
            mb_df = aggregate_rows(7, 26)
            cw_df = aggregate_rows(27, 44)

            # Check if output file exists to append or create new
            if os.path.exists(out_path):
                try: 
                    wb = load_workbook(out_path)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not load existing file: {str(e)}")
                    return
            else:
                wb = Workbook()

            # Helper to generate unique sheet name
            def get_u_name(base):
                cand = base
                cnt = 1
                while cand in wb.sheetnames:
                    cand = (base[:26] + f" ({cnt})")[:31]
                    cnt += 1
                return cand
            
            db_title = get_u_name("DASHBOARD")
            mb_title = get_u_name("MB Detail")
            cw_title = get_u_name("CW Detail")
            ce_title = get_u_name("CALC_ENGINE")

            # Create or identify active sheet
            if db_title == "DASHBOARD" and "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                ws = wb.active
                ws.title = db_title
            else:
                ws = wb.create_sheet(db_title)
            
            ws.sheet_view.showGridLines = False
            
            # Create detail sheets
            ws_mb_detail = wb.create_sheet(mb_title)
            ws_cw_detail = wb.create_sheet(cw_title)
            
            # Setup Canvas for Dashboard
            for r in range(1, 45): # Limit range for dashboard
                for c in range(1, 40):
                    ws.cell(row=r, column=c).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Header
            head_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
            ws.merge_cells("B2:AH5")
            cell_h = ws["B2"]
            cell_h.value = "JKD CONSTRUCTION ANALYTICS DASHBOARD"
            cell_h.font = Font(bold=True, size=26, color="FFFFFF")
            cell_h.fill = head_fill
            cell_h.alignment = Alignment(horizontal='center', vertical='center')
            
            period_str = f"PERIOD: {start_d.strftime('%d %b %Y')} to {end_d.strftime('%d %b %Y')}"
            ws.merge_cells("B6:AH6")
            ws["B6"].value = period_str.upper()
            ws["B6"].font = Font(bold=True, size=11, color="64748B")
            ws["B6"].alignment = Alignment(horizontal='center')
            
            # Chart Data Sheet (CALC ENGINE)
            ws_cat = wb.create_sheet(ce_title)
            ws_cat.sheet_state = 'hidden'

            # MB Chart Data - LINKED TO DETAIL SHEET
            ws_cat["A1"] = "Activity"; ws_cat["B1"] = "Planned"; ws_cat["C1"] = "Achieved"
            for i, r in mb_df.iterrows():
                row_idx = i + 2
                detail_row = i + 6
                ws_cat.cell(row=row_idx, column=1, value=f"='{mb_title}'!A{detail_row}")
                ws_cat.cell(row=row_idx, column=2, value=f"='{mb_title}'!C{detail_row}")
                ws_cat.cell(row=row_idx, column=3, value=f"='{mb_title}'!D{detail_row}")
            
            # CW Chart Data - LINKED TO DETAIL SHEET
            ws_cat["E1"] = "Activity"; ws_cat["F1"] = "Planned"; ws_cat["G1"] = "Achieved"
            for i, r in cw_df.iterrows():
                row_idx = i + 2
                detail_row = i + 6
                ws_cat.cell(row=row_idx, column=5, value=f"='{cw_title}'!A{detail_row}")
                ws_cat.cell(row=row_idx, column=6, value=f"='{cw_title}'!C{detail_row}")
                ws_cat.cell(row=row_idx, column=7, value=f"='{cw_title}'!D{detail_row}")

            # Summary for KPI - DYNAMIC FORMULAS
            mb_max = len(mb_df) + 1
            cw_max = len(cw_df) + 1
            ws_cat["I1"] = "Total Planned"; ws_cat["I2"] = f"=SUM(B2:B{mb_max}) + SUM(F2:F{cw_max})"
            ws_cat["J1"] = "Total Achieved"; ws_cat["J2"] = f"=SUM(C2:C{mb_max}) + SUM(G2:G{cw_max})"
            ws_cat["K1"] = "Variance"; ws_cat["K2"] = "=J2-I2"
            ws_cat["L1"] = "Achievement %"; ws_cat["L2"] = "=IF(I2=0,0,J2/I2)"

            # KPI Row - LINKED TO CALC_ENGINE
            draw_kpi_card(ws, 9, 3, "TOTAL PLANNED", f"='{ce_title}'!I2", "Aggregated Units", "f1f5f9", num_format="#,##0.0")
            draw_kpi_card(ws, 9, 8, "TOTAL ACHIEVED", f"='{ce_title}'!J2", "Aggregated Units", "f1f5f9", num_format="#,##0.0")
            draw_kpi_card(ws, 9, 13, "VARIANCE", f"='{ce_title}'!K2", "Behind/Ahead", "f1f5f9", num_format="#,##0.0")
            draw_kpi_card(ws, 9, 18, "ACHIEVEMENT %", f"='{ce_title}'!L2", "Overall Efficiency", "f1f5f9", num_format="0.0%")

            # Sections
            ws.merge_cells("B15:S15")
            ws["B15"].value = "MAIN FACTORY BUILDING DETAILS"
            ws["B15"].font = Font(bold=True, size=14, color="1e40af")
            
            ws.merge_cells("U15:AH15")
            ws["U15"].value = "COMPOUND WALL DETAILS"
            ws["U15"].font = Font(bold=True, size=14, color="1e40af")

            # Charts - Spanning vertically to prevent overlap
            if not mb_df.empty:
                chart_mb = create_premium_chart(ws_cat, "Main Building Progress", 
                                              {'min_col': 2, 'max_col': 3, 'min_row': 1, 'max_row': len(mb_df)+1},
                                              {'min_col': 1, 'min_row': 2, 'max_row': len(mb_df)+1})
                ws.add_chart(chart_mb, "B17")

            if not cw_df.empty:
                chart_cw = create_premium_chart(ws_cat, "Compound Wall Progress", 
                                              {'min_col': 6, 'max_col': 7, 'min_row': 1, 'max_row': len(cw_df)+1},
                                              {'min_col': 5, 'min_row': 2, 'max_row': len(cw_df)+1})
                # Moved much lower to prevent overlap with the first chart
                ws.add_chart(chart_cw, "B55")

            # Data Tables - Now in separate sheets
            def write_premium_table(target_ws, title, dataframe):
                target_ws.sheet_view.showGridLines = False
                
                # Sheet Header
                target_ws.merge_cells("A1:F2")
                h_cell = target_ws["A1"]
                h_cell.value = f"{title.upper()} - DETAILED PROGRESS"
                h_cell.font = Font(bold=True, size=16, color="FFFFFF")
                h_cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
                h_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                target_ws.merge_cells("A3:F3")
                target_ws["A3"].value = f"ANALYSIS PERIOD: {start_d.strftime('%d %b %Y')} TO {end_d.strftime('%d %b %Y')}"
                target_ws["A3"].font = Font(italic=True, color="64748B")
                target_ws["A3"].alignment = Alignment(horizontal='center')

                heads = ["Activity", "Unit", "Planned Qty", "Achieved Qty", "Variation", "Achievement %"]
                for i, h in enumerate(heads):
                    c = target_ws.cell(row=5, column=i+1, value=h)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
                    c.alignment = Alignment(horizontal='center')
                    c.border = Border(left=Side(style='thin', color="FFFFFF"), right=Side(style='thin', color="FFFFFF"))
                
                curr_row = 6
                stripe_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
                
                for i, (_, r) in enumerate(dataframe.iterrows()):
                    # Zebra striping
                    row_fill = stripe_fill if i % 2 == 1 else None
                    
                    c1 = target_ws.cell(row=curr_row, column=1, value=r['Activity'])
                    c1.border = Border(bottom=Side(style='thin', color="cbd5e1"))
                    if row_fill: c1.fill = row_fill
                    
                    c2 = target_ws.cell(row=curr_row, column=2, value=r['Unit'])
                    c2.alignment = Alignment(horizontal='center')
                    if row_fill: c2.fill = row_fill
                    
                    c3 = target_ws.cell(row=curr_row, column=3, value=r['Planned'])
                    if row_fill: c3.fill = row_fill
                    
                    c4 = target_ws.cell(row=curr_row, column=4, value=r['Achieved'])
                    if row_fill: c4.fill = row_fill
                    
                    # VARIATION FORMULA
                    v_cell = target_ws.cell(row=curr_row, column=5, value=f"=D{curr_row}-C{curr_row}")
                    v_cell.font = Font(bold=True)
                    if row_fill: v_cell.fill = row_fill
                    
                    # ACHIEVEMENT % FORMULA
                    p_cell = target_ws.cell(row=curr_row, column=6, value=f"=IF(C{curr_row}=0,0,D{curr_row}/C{curr_row})")
                    p_cell.number_format = '0.0%'
                    if row_fill: p_cell.fill = row_fill
                    
                    curr_row += 1
                
                # Apply Conditional Formatting
                red_fill = PatternFill(start_color="fee2e2", end_color="991b1b", fill_type="solid") # Red bg, dark text? actually just bg
                green_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
                yellow_fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")
                
                # Variation CF
                target_ws.conditional_formatting.add(f'E6:E{curr_row-1}',
                    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
                target_ws.conditional_formatting.add(f'E6:E{curr_row-1}',
                    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
                
                # Achievement CF
                target_ws.conditional_formatting.add(f'F6:F{curr_row-1}',
                    CellIsRule(operator='greaterThanOrEqual', formula=['1'], fill=green_fill))
                target_ws.conditional_formatting.add(f'F6:F{curr_row-1}',
                    CellIsRule(operator='between', formula=['0.8', '0.99'], fill=yellow_fill))
                target_ws.conditional_formatting.add(f'F6:F{curr_row-1}',
                    CellIsRule(operator='lessThan', formula=['0.8'], fill=red_fill))
                
                target_ws.column_dimensions['A'].width = 45
                target_ws.column_dimensions['B'].width = 12
                target_ws.column_dimensions['C'].width = 15
                target_ws.column_dimensions['D'].width = 15
                target_ws.column_dimensions['E'].width = 15
                target_ws.column_dimensions['F'].width = 18

            if not mb_df.empty:
                write_premium_table(ws_mb_detail, "Main Factory Building", mb_df)
            if not cw_df.empty:
                write_premium_table(ws_cw_detail, "Compound Wall", cw_df)

            # Finalize Dashboard sizing
            for i in range(1, 40): ws.column_dimensions[get_column_letter(i)].width = 3.5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['U'].width = 30

            wb.save(out_path)
            messagebox.showinfo("Success", "Your Unimaginable Dashboard is Ready!")
            os.startfile(out_path)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed: {str(e)}")
        finally:
            self.gen_btn.config(state="normal", text="GENERATE UNIMAGINABLE DASHBOARD")

if __name__ == "__main__":
    root = tk.Tk()
    app = DashingDashboardApp(root)
    root.mainloop()
