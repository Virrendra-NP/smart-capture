import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_excel(r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\JKD - DPR - 28.02.2026 (1) (1) (1).xlsx', sheet_name="Feb'26", header=None)

date_cols = {41: '21-Feb', 43: '23-Feb', 45: '24-Feb', 47: '25-Feb', 49: '26-Feb', 51: '27-Feb', 53: '28-Feb'}

def get_weekly_data(start_row, end_row):
    data_rows = []
    for i in range(start_row, end_row):
        activity = df.iloc[i, 1]
        unit = df.iloc[i, 2]
        if pd.notna(activity) and activity not in ['Main Factory Building', 'Compound Wall'] and not isinstance(activity, float):
            p, a = 0, 0
            daily_data = {}
            for col in date_cols.keys():
                planned_val = float(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else 0
                achieved_val = float(df.iloc[i, col + 1]) if pd.notna(df.iloc[i, col + 1]) else 0
                p += planned_val
                a += achieved_val
                daily_data[col] = {'planned': planned_val, 'achieved': achieved_val}
            if p > 0 or a > 0:
                data_rows.append({'Activity': activity, 'Unit': unit, 'Planned': p, 'Achieved': a, '%': round((a/p*100) if p>0 else 0, 1), 'daily': daily_data})
    return pd.DataFrame(data_rows)

main_building = get_weekly_data(7, 26)
compound_wall = get_weekly_data(27, 40)

wb = Workbook()
ws_data = wb.active
ws_data.title = "Data"

ws_data['A1'] = "JKD 358 - Weekly Progress Report (21-Feb-2026 to 28-Feb-2026)"
ws_data['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws_data['A1'].fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
ws_data.merge_cells('A1:E1')
ws_data['A1'].alignment = Alignment(horizontal='center')

ws_data['A3'] = "Category"
ws_data['B3'] = "Activity"
ws_data['C3'] = "Unit"
ws_data['D3'] = "Planned"
ws_data['E3'] = "Achieved"
ws_data['F3'] = "Achievement %"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_data[f'{col}3'].font = Font(bold=True, color='FFFFFF')
    ws_data[f'{col}3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_data[f'{col}3'].alignment = Alignment(horizontal='center')

row = 4
mb_start = row
ws_data[f'A{row}'] = "MAIN FACTORY BUILDING"
ws_data[f'A{row}'].font = Font(bold=True)
ws_data.merge_cells(f'A{row}:F{row}')
ws_data[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
row += 1

for _, r in main_building.iterrows():
    ws_data[f'A{row}'] = "Main Factory Building"
    ws_data[f'B{row}'] = r['Activity']
    ws_data[f'C{row}'] = r['Unit']
    ws_data[f'D{row}'] = r['Planned']
    ws_data[f'E{row}'] = r['Achieved']
    ws_data[f'F{row}'] = r['%']
    row += 1

mb_end = row - 1
ws_data[f'A{row}'] = "Main Factory Building"
ws_data[f'B{row}'] = "TOTAL"
ws_data[f'D{row}'] = main_building['Planned'].sum()
ws_data[f'E{row}'] = main_building['Achieved'].sum()
ws_data[f'F{row}'] = round(main_building['Achieved'].sum()/main_building['Planned'].sum()*100, 1) if main_building['Planned'].sum()>0 else 0
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_data[f'{col}{row}'].font = Font(bold=True)
row += 2

cw_start = row
ws_data[f'A{row}'] = "COMPOUND WALL"
ws_data[f'A{row}'].font = Font(bold=True)
ws_data.merge_cells(f'A{row}:F{row}')
ws_data[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
row += 1

for _, r in compound_wall.iterrows():
    ws_data[f'A{row}'] = "Compound Wall"
    ws_data[f'B{row}'] = r['Activity']
    ws_data[f'C{row}'] = r['Unit']
    ws_data[f'D{row}'] = r['Planned']
    ws_data[f'E{row}'] = r['Achieved']
    ws_data[f'F{row}'] = r['%']
    row += 1

cw_end = row - 1
ws_data[f'A{row}'] = "Compound Wall"
ws_data[f'B{row}'] = "TOTAL"
ws_data[f'D{row}'] = compound_wall['Planned'].sum()
ws_data[f'E{row}'] = compound_wall['Achieved'].sum()
ws_data[f'F{row}'] = round(compound_wall['Achieved'].sum()/compound_wall['Planned'].sum()*100, 1) if compound_wall['Planned'].sum()>0 else 0
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_data[f'{col}{row}'].font = Font(bold=True)
row += 2

grand_total_row = row
ws_data[f'A{row}'] = "GRAND TOTAL"
ws_data[f'B{row}'] = ""
ws_data[f'D{row}'] = main_building['Planned'].sum() + compound_wall['Planned'].sum()
ws_data[f'E{row}'] = main_building['Achieved'].sum() + compound_wall['Achieved'].sum()
ws_data[f'F{row}'] = round((main_building['Achieved'].sum() + compound_wall['Achieved'].sum())/(main_building['Planned'].sum() + compound_wall['Planned'].sum())*100, 1)
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_data[f'{col}{row}'].font = Font(bold=True, size=12)
    ws_data[f'{col}{row}'].fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')

ws_data.column_dimensions['A'].width = 22
ws_data.column_dimensions['B'].width = 35
ws_data.column_dimensions['C'].width = 10
ws_data.column_dimensions['D'].width = 12
ws_data.column_dimensions['E'].width = 12
ws_data.column_dimensions['F'].width = 15

mb_planned = main_building['Planned'].sum()
mb_achieved = main_building['Achieved'].sum()
cw_planned = compound_wall['Planned'].sum()
cw_achieved = compound_wall['Achieved'].sum()
total_planned = mb_planned + cw_planned
total_achieved = mb_achieved + cw_achieved

ws_dash = wb.create_sheet("KPI Dashboard")

ws_dash['A1'] = "KPI DASHBOARD - JKD 358 Production Block"
ws_dash['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws_dash['A1'].fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
ws_dash.merge_cells('A1:H1')
ws_dash['A1'].alignment = Alignment(horizontal='center')
ws_dash.row_dimensions[1].height = 25

ws_dash['A2'] = "Weekly Progress: 21-Feb-2026 to 28-Feb-2026"
ws_dash['A2'].font = Font(size=11, italic=True)
ws_dash.merge_cells('A2:H2')
ws_dash['A2'].alignment = Alignment(horizontal='center')

kpi_data_start = 4
kpis = [
    ['KPI Metric', 'Value'],
    ['Total Planned', total_planned],
    ['Total Achieved', total_achieved],
    ['Overall Achievement %', round(total_achieved/total_planned*100, 1)],
    ['Main Building Planned', mb_planned],
    ['Main Building Achieved', mb_achieved],
    ['Main Building Achievement %', round(mb_achieved/mb_planned*100, 1) if mb_planned>0 else 0],
    ['Compound Wall Planned', cw_planned],
    ['Compound Wall Achieved', cw_achieved],
    ['Compound Wall Achievement %', round(cw_achieved/cw_planned*100, 1) if cw_planned>0 else 0],
]

for i, (label, value) in enumerate(kpis, kpi_data_start):
    ws_dash.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
    ws_dash.cell(row=i, column=2, value=value).font = Font(size=14, bold=True, color='2F5496')
    ws_dash.cell(row=i, column=1).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    ws_dash.cell(row=i, column=2).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

ws_dash.column_dimensions['A'].width = 28
ws_dash.column_dimensions['B'].width = 18

chart_data_start = 18
ws_dash.cell(row=chart_data_start, column=1, value="Section")
ws_dash.cell(row=chart_data_start, column=2, value="Planned")
ws_dash.cell(row=chart_data_start, column=3, value="Achieved")

ws_dash.cell(row=chart_data_start+1, column=1, value="Main Factory Building")
ws_dash.cell(row=chart_data_start+1, column=2, value=mb_planned)
ws_dash.cell(row=chart_data_start+1, column=3, value=mb_achieved)

ws_dash.cell(row=chart_data_start+2, column=1, value="Compound Wall")
ws_dash.cell(row=chart_data_start+2, column=2, value=cw_planned)
ws_dash.cell(row=chart_data_start+2, column=3, value=cw_achieved)

ws_dash.cell(row=chart_data_start+3, column=1, value="Total")
ws_dash.cell(row=chart_data_start+3, column=2, value=total_planned)
ws_dash.cell(row=chart_data_start+3, column=3, value=total_achieved)

for col in range(1, 4):
    ws_dash.cell(row=chart_data_start, column=col).font = Font(bold=True, color='FFFFFF')
    ws_dash.cell(row=chart_data_start, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_dash.cell(row=chart_data_start, column=col).alignment = Alignment(horizontal='center')
    ws_dash.cell(row=chart_data_start+3, column=col).font = Font(bold=True)
    ws_dash.cell(row=chart_data_start+3, column=col).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Planned vs Achieved by Section"
chart1.y_axis.title = "Quantity (Units)"
chart1.x_axis.title = "Section"

data = Reference(ws_dash, min_col=2, min_row=chart_data_start, max_row=chart_data_start+3, max_col=3)
cats = Reference(ws_dash, min_col=1, min_row=chart_data_start+1, max_row=chart_data_start+3)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
chart1.width = 14
chart1.height = 10

ws_dash.add_chart(chart1, "E4")

pie_data_start = 18
ws_dash.cell(row=pie_data_start, column=5, value="Progress")
ws_dash.cell(row=pie_data_start+1, column=5, value="Achieved")
ws_dash.cell(row=pie_data_start+2, column=5, value="Pending")
ws_dash.cell(row=pie_data_start+1, column=6, value=total_achieved)
ws_dash.cell(row=pie_data_start+2, column=6, value=total_planned - total_achieved)

for col in range(5, 7):
    ws_dash.cell(row=pie_data_start, column=col).font = Font(bold=True, color='FFFFFF')
    ws_dash.cell(row=pie_data_start, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

pie = PieChart()
pie.title = "Overall Progress"
labels = Reference(ws_dash, min_col=5, min_row=pie_data_start+1, max_row=pie_data_start+2)
data = Reference(ws_dash, min_col=6, min_row=pie_data_start, max_row=pie_data_start+2)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 12
pie.height = 10

ws_dash.add_chart(pie, "E20")

activity_chart_start = 35
ws_dash.cell(row=activity_chart_start, column=1, value="Activity")
ws_dash.cell(row=activity_chart_start, column=2, value="Planned")
ws_dash.cell(row=activity_chart_start, column=3, value="Achieved")
ws_dash.cell(row=activity_chart_start, column=4, value="Achievement %")

all_activities = pd.concat([main_building, compound_wall]).sort_values('%', ascending=False)

for idx, (_, r) in enumerate(all_activities.iterrows(), activity_chart_start+1):
    ws_dash.cell(row=idx, column=1, value=r['Activity'][:30])
    ws_dash.cell(row=idx, column=2, value=r['Planned'])
    ws_dash.cell(row=idx, column=3, value=r['Achieved'])
    ws_dash.cell(row=idx, column=4, value=r['%'])

for col in range(1, 5):
    ws_dash.cell(row=activity_chart_start, column=col).font = Font(bold=True, color='FFFFFF')
    ws_dash.cell(row=activity_chart_start, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_dash.cell(row=activity_chart_start, column=col).alignment = Alignment(horizontal='center')

chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Achievement % by Activity"
chart2.y_axis.title = "Activity"
chart2.x_axis.title = "Achievement %"

data = Reference(ws_dash, min_col=4, min_row=activity_chart_start, max_row=activity_chart_start + len(all_activities))
cats = Reference(ws_dash, min_col=1, min_row=activity_chart_start+1, max_row=activity_chart_start + len(all_activities))
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.width = 16
chart2.height = 12

ws_dash.add_chart(chart2, "A45")

output_path = r'D:\JKD_Folder\JKD-PROJECT SITE\DPR\Excel_DPR\Weekly_Progress_Report_21-28_Feb_2026.xlsx'
wb.save(output_path)
print(f"Excel saved: {output_path}")
print(f"\n=== SUMMARY ===")
print(f"Main Building - Planned: {mb_planned:.2f}, Achieved: {mb_achieved:.2f}, %: {round(mb_achieved/mb_planned*100, 1)}%")
print(f"Compound Wall - Planned: {cw_planned:.2f}, Achieved: {cw_achieved:.2f}, %: {round(cw_achieved/cw_planned*100, 1)}%")
print(f"Total - Planned: {total_planned:.2f}, Achieved: {total_achieved:.2f}, %: {round(total_achieved/total_planned*100, 1)}%")
